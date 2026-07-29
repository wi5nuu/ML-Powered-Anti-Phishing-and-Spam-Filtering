"""
Admin routes — RBAC-protected admin API endpoints.

Registered via app.include_router(admin_routes.router) in app.py.

Routes here are CRUD operations not covered (or covered with less validation)
by the @app.xxx routes in app.py. Specifically:
  - POST /users  — Pydantic-validated creation with proper role guards
  - PATCH /users/id/{user_id}  — partial update by numeric id
  - DELETE /users/id/{user_id}  — delete by numeric id (superadmin only)
  - PATCH /mailboxes/id/{mailbox_id}  — partial update
  - DELETE /mailboxes/id/{mailbox_id}  — delete (superadmin only, stricter than app.py)

Intentionally NOT duplicated here (app.py versions are richer):
  GET  /users      — org-scoped filtering
  GET  /mailboxes  — returns forward_to, assigned_to, storage_bytes, etc.
  POST /mailboxes  — rate-limited, handles password + assigned_to
  GET  /stats      — app.py version aggregates email/threat counts too
"""

from fastapi import APIRouter, Depends, HTTPException, Request, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional, List
import datetime
import csv
import io
import json
import os
from zoneinfo import ZoneInfo
from email.utils import getaddresses
from html import escape

from database.models import User, UserRole, AdminMailbox, AdminMailboxAccess, AuditLog, Organization, QuarantineEmail
from dashboard.database import get_db
from dashboard.auth import get_current_user_cookie, hash_password, log_audit
from dashboard.config import get_configured_mail_domain, email_uses_configured_domain, is_valid_email_address
from sqlalchemy import DateTime, String, case, cast, func, or_

# Import libraries for PDF and Excel export
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

router = APIRouter(prefix="/api/admin", tags=["admin-crud"])


# ── RBAC dependency classes ────────────────────────────────────────────────────
# FastAPI can inject Request directly into Depends() functions — the guard
# functions below receive both Request and Session via FastAPI's DI system.

def require_admin(request: Request, db: Session = Depends(get_db)) -> User:
    """Dependency: requires admin or superadmin role, authenticated via cookie."""
    user = get_current_user_cookie(request, db)
    if user.role not in (UserRole.ADMIN.value, UserRole.SUPERADMIN.value):
        raise HTTPException(status_code=403, detail="Admin access required")
    return user


def require_superadmin(request: Request, db: Session = Depends(get_db)) -> User:
    """Dependency: requires superadmin role, authenticated via cookie."""
    user = get_current_user_cookie(request, db)
    if user.role != UserRole.SUPERADMIN.value:
        raise HTTPException(status_code=403, detail="Superadmin access required")
    return user


def _admin_can_manage_user(admin: User, user: User) -> bool:
    if admin.role == UserRole.SUPERADMIN.value:
        return True
    if user.role != UserRole.USER.value:
        return False
    if admin.organization_id:
        return user.organization_id == admin.organization_id
    return user.organization_id is None and email_uses_configured_domain(user.email or "")


# ── Pydantic schemas ───────────────────────────────────────────────────────────

class UserCreate(BaseModel):
    username: str
    email: Optional[str] = None
    password: str
    role: str = "user"
    is_active: bool = True
    organization_id: Optional[int] = None


class UserUpdate(BaseModel):
    email: Optional[str] = None
    role: Optional[str] = None
    is_active: Optional[bool] = None
    password: Optional[str] = None


class MailboxUpdate(BaseModel):
    sender_name: Optional[str] = None
    is_active: Optional[bool] = None


@router.get("/config")
def get_admin_config(current_user: User = Depends(require_admin)):
    """Expose non-secret deployment values needed by admin forms."""
    return {"mail_domain": get_configured_mail_domain()}


# ── Helpers ────────────────────────────────────────────────────────────────────

def _serialize_user(u: User) -> dict:
    data = {
        "id": u.id,
        "username": u.username,
        "role": u.role,
        "is_active": getattr(u, "is_active", True),
        "created_at": str(getattr(u, "created_at", "")),
    }
    if u.role == UserRole.USER.value:
        data["email"] = getattr(u, "email", "") or ""
    return data


def _serialize_mailbox(m: AdminMailbox) -> dict:
    return {
        "id": m.id,
        "email": m.email,
        "domain": getattr(m, "domain", ""),
        "sender_name": getattr(m, "sender_name", ""),
        "avatar_url": getattr(m, "avatar_url", "") or "",
        "assigned_to": getattr(m, "assigned_to", "") or "",
        "created_by": getattr(m, "created_by", ""),
        "is_active": getattr(m, "is_active", True),
        "created_at": str(getattr(m, "created_at", "")),
    }


# ── User endpoints ─────────────────────────────────────────────────────────────

@router.post("/users", status_code=201)
def create_user(
    request: Request,
    payload: UserCreate,
    current_user: User = Depends(require_superadmin),
    db: Session = Depends(get_db),
):
    """Create an administrator account from the superadmin dashboard."""
    username = payload.username.strip()
    if not username:
        raise HTTPException(status_code=400, detail="Username wajib diisi.")
    if len(username) > 64:
        raise HTTPException(status_code=400, detail="Username maksimal 64 karakter.")
    if len(payload.password) < 8:
        raise HTTPException(status_code=400, detail="Password minimal 8 karakter.")
    if payload.role != UserRole.ADMIN.value:
        raise HTTPException(status_code=400, detail="Halaman ini hanya dapat membuat akun admin.")
    if db.query(User).filter(User.username == username).first():
        raise HTTPException(status_code=409, detail="Username sudah digunakan.")
    
    # Organization is optional for global admins. When provided, it must exist.
    org_id = payload.organization_id
    if org_id and not db.query(Organization).filter(Organization.id == org_id).first():
        raise HTTPException(status_code=400, detail="Organisasi tidak ditemukan.")
    
    user = User(
        username=username,
        email=None,
        hashed_password=hash_password(payload.password),
        role=payload.role,
        is_active=payload.is_active,
        organization_id=org_id,
        created_at=datetime.datetime.now(datetime.timezone.utc),
    )
    db.add(user)
    log_audit(
        db,
        current_user.username,
        "create_user",
        None,
        request.client.host if request.client else None,
        username,
    )
    db.commit()
    db.refresh(user)
    return _serialize_user(user)


@router.patch("/users/id/{user_id}")
def update_user(
    user_id: int,
    payload: UserUpdate,
    current_user: User = Depends(require_superadmin),
    db: Session = Depends(get_db),
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Pengguna tidak ditemukan.")
    if user.role != UserRole.ADMIN.value:
        raise HTTPException(status_code=403, detail="Halaman ini hanya dapat mengelola akun admin.")
    if payload.role == UserRole.SUPERADMIN.value and current_user.role != UserRole.SUPERADMIN.value:
        raise HTTPException(status_code=403, detail="Hanya superadmin yang dapat menetapkan role superadmin.")
    if user.id == current_user.id and payload.is_active is False:
        raise HTTPException(status_code=400, detail="Tidak dapat menonaktifkan akun Anda sendiri.")
    # Admins may edit only users in their organization.
    if current_user.role == UserRole.ADMIN.value:
        if not _admin_can_manage_user(current_user, user):
            raise HTTPException(status_code=403, detail="Admin hanya dapat mengedit pengguna dalam cakupannya.")
        if user.role != UserRole.USER.value:
            raise HTTPException(status_code=403, detail="Admin hanya dapat mengedit pengguna dengan role 'user'.")
        if payload.role is not None and payload.role != UserRole.USER.value:
            raise HTTPException(status_code=403, detail="Admin hanya dapat menetapkan role 'user'.")
        if payload.email is not None and not email_uses_configured_domain(payload.email):
            raise HTTPException(
                status_code=400,
                detail=f"Email pengguna wajib menggunakan domain @{get_configured_mail_domain()}.",
            )
    user.email = None
    if payload.role is not None:
        user.role = payload.role
    if payload.is_active is not None:
        user.is_active = payload.is_active
    if payload.password:
        user.hashed_password = hash_password(payload.password)
    db.commit()
    db.refresh(user)
    return _serialize_user(user)


@router.delete("/users/id/{user_id}", status_code=204)
def delete_user(
    user_id: int,
    current_user: User = Depends(require_superadmin),
    db: Session = Depends(get_db),
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Pengguna tidak ditemukan.")
    if user.role != UserRole.ADMIN.value:
        raise HTTPException(status_code=403, detail="Halaman ini hanya dapat menghapus akun admin.")
    if user.id == current_user.id:
        raise HTTPException(status_code=400, detail="Tidak dapat menghapus akun Anda sendiri.")
    owned_mailboxes = db.query(AdminMailbox).filter(
        AdminMailbox.assigned_to == user.username,
        AdminMailbox.is_active == True,
    ).count()
    if owned_mailboxes:
        raise HTTPException(
            status_code=409,
            detail=f"Pindahkan {owned_mailboxes} mailbox yang dikelola admin ini sebelum menghapus akun.",
        )
    if user.role == UserRole.SUPERADMIN.value:
        remaining_superadmins = db.query(User).filter(
            User.role == UserRole.SUPERADMIN.value,
            User.id != user.id,
            User.is_active == True,
        ).count()
        if remaining_superadmins == 0:
            raise HTTPException(status_code=400, detail="Minimal satu superadmin aktif harus tetap tersedia.")
    db.query(AdminMailboxAccess).filter(
        AdminMailboxAccess.username == user.username
    ).delete(synchronize_session=False)
    db.delete(user)
    db.commit()


# ── Mailbox endpoints ──────────────────────────────────────────────────────────

@router.patch("/mailboxes/id/{mailbox_id}")
def update_mailbox(
    mailbox_id: int,
    payload: MailboxUpdate,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    mailbox = db.query(AdminMailbox).filter(AdminMailbox.id == mailbox_id).first()
    if not mailbox:
        raise HTTPException(status_code=404, detail="Mailbox tidak ditemukan.")
    if current_user.role == UserRole.ADMIN.value:
        if mailbox.assigned_to != current_user.username:
            raise HTTPException(status_code=403, detail="Admin tidak memiliki akses ke mailbox ini.")
    if payload.sender_name is not None:
        mailbox.sender_name = payload.sender_name
    if payload.is_active is not None:
        mailbox.is_active = payload.is_active
    db.commit()
    db.refresh(mailbox)
    return _serialize_mailbox(mailbox)


@router.delete("/mailboxes/id/{mailbox_id}", status_code=204)
def delete_mailbox(
    mailbox_id: int,
    current_user: User = Depends(require_superadmin),
    db: Session = Depends(get_db),
):
    """Delete a mailbox. Requires superadmin (stricter than app.py's admin-level delete)."""
    mailbox = db.query(AdminMailbox).filter(AdminMailbox.id == mailbox_id).first()
    if not mailbox:
        raise HTTPException(status_code=404, detail="Mailbox tidak ditemukan.")
    db.delete(mailbox)
    db.commit()


# ── Audit Trail Export endpoint ────────────────────────────────────────────


def _generate_csv_export(audit_records):
    """Generate CSV export of audit log data."""
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow(["ID","Timestamp","User","Action","Email ID","Details"])
    for record in audit_records:
        writer.writerow([
            record.id,
            record.created_at.strftime("%Y-%m-%d %H:%M:%S") if record.created_at else "",
            record.user or "", record.action or "", record.email_id or "", record.details or ""
        ])
    output.seek(0)
    return output.getvalue()


def _generate_excel_export(audit_records):
    """Generate Excel export of audit log data."""
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=500, detail="openpyxl library not installed. Run: pip install openpyxl")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Log"
    hdr_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_align = Alignment(horizontal="center", vertical="center")
    ws.append(["ID","Timestamp","User","Action","Email ID","Details"])
    for cell in ws[1]:
        cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = hdr_align
    for record in audit_records:
        ws.append([
            record.id,
            record.created_at.strftime("%Y-%m-%d %H:%M:%S") if record.created_at else "",
            record.user or "", record.action or "", record.email_id or "", record.details or ""
        ])
    for col in ws.columns:
        letter = col[0].column_letter
        mx = min(max((len(str(c.value)) for c in col if c.value), default=0) + 2, 50)
        ws.column_dimensions[letter].width = mx
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out.getvalue()


def _generate_pdf_export(audit_records):
    """Generate PDF export of audit log data."""
    if not REPORTLAB_AVAILABLE:
        raise HTTPException(status_code=500, detail="reportlab library not installed. Run: pip install reportlab")
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    el = []
    styles = getSampleStyleSheet()
    ts = ParagraphStyle('T',parent=styles['Heading1'],fontSize=18,
                         textColor=colors.HexColor('#1a73e8'),spaceAfter=30,alignment=1)
    el.append(Paragraph("User Activity Tracking Report", ts))
    el.append(Paragraph(f"Generated: {datetime.datetime.now(datetime.timezone.utc).strftime('%Y-%m-%d %H:%M:%S')} UTC", styles['Normal']))
    el.append(Spacer(1,20))
    td = [["ID","Timestamp","User","Action","Email ID"]]
    for r in audit_records:
        td.append([str(r.id),
                   r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "",
                   r.user or "", r.action or "",
                   str(r.email_id) if r.email_id else ""])
    t = Table(td, colWidths=[0.6*inch,1.5*inch,1.2*inch,1.2*inch,1*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1a73e8')),
        ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,0),10),
        ('BOTTOMPADDING',(0,0),(-1,0),12),
        ('GRID',(0,0),(-1,-1),0.5,colors.grey),
        ('FONTSIZE',(0,1),(-1,-1),8),
        ('VALIGN',(0,0),(-1,-1),'TOP'),
    ]))
    el.append(t)
    doc.build(el); buf.seek(0)
    return buf.getvalue()


@router.get("/track/export")
def export_audit_trail(
    format: str = Query("csv", description="Export format: csv, excel, or pdf"),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Export audit log data (legacy)."""
    # Export exactly the same persisted audit records the caller is allowed to
    # view.  Admins are restricted to their organization; superadmins retain
    # the global scope.  No synthetic rows are created during export.
    audit_query = db.query(AuditLog)
    if current_user.role != UserRole.SUPERADMIN.value:
        if current_user.organization_id:
            org_usernames = [
                username for (username,) in db.query(User.username).filter(
                    User.organization_id == current_user.organization_id
                ).all()
            ]
            audit_query = audit_query.filter(AuditLog.user.in_(org_usernames))
        else:
            # Match the read endpoint: an admin without an organization has no
            # organization-scoped audit records to export.
            audit_query = audit_query.filter(False)
    audit_records = audit_query.order_by(AuditLog.created_at.desc()).limit(500).all()
    ts = datetime.datetime.now(datetime.timezone.utc).strftime("%Y%m%d_%H%M%S")
    fl = format.lower()
    if fl == "csv":
        content = _generate_csv_export(audit_records)
        mt = "text/csv"; fn = f"audit_log_{ts}.csv"
    elif fl == "excel":
        content = _generate_excel_export(audit_records)
        mt = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"; fn = f"audit_log_{ts}.xlsx"
    elif fl == "pdf":
        content = _generate_pdf_export(audit_records)
        mt = "application/pdf"; fn = f"audit_log_{ts}.pdf"
    else:
        raise HTTPException(status_code=400, detail="Unsupported format. Use csv, excel, or pdf.")
    return StreamingResponse(iter([content]), media_type=mt,
                             headers={"Content-Disposition": f"attachment; filename={fn}"})


# ── Comprehensive Export (Report) ──────────────────────────────────────────


class ExportRequest(BaseModel):
    format: str = "pdf"
    date_from: Optional[str] = None
    date_to: Optional[str] = None
    admin_ids: Optional[List[int]] = None
    mailbox_ids: Optional[List[int]] = None
    include_users: bool = True
    include_emails: bool = True


def _report_timezone() -> ZoneInfo:
    try:
        return ZoneInfo(os.getenv("APP_TIMEZONE", "Asia/Jakarta"))
    except Exception:
        return ZoneInfo("UTC")


def _parse_date(s: Optional[str]) -> Optional[datetime.datetime]:
    if not s:
        return None
    try:
        local_date = datetime.datetime.strptime(s, "%Y-%m-%d").date()
        local_midnight = datetime.datetime.combine(
            local_date,
            datetime.time.min,
            tzinfo=_report_timezone(),
        )
        return local_midnight.astimezone(datetime.timezone.utc)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail=f"Invalid date: {s}. Use YYYY-MM-DD.")


def _report_received_at_expression(db: Session):
    """Return a comparable timestamp for both current and legacy schemas.

    Some deployed databases predate the SQLAlchemy DateTime declaration and
    still store ``received_at`` as VARCHAR. PostgreSQL cannot compare that
    column directly with timezone-aware report boundaries. Cast only strings
    that begin with an ISO date so a malformed legacy value does not make the
    complete report fail.
    """
    dialect = getattr(getattr(db, "bind", None), "dialect", None)
    if getattr(dialect, "name", "") != "postgresql":
        return QuarantineEmail.received_at
    received_text = cast(QuarantineEmail.received_at, String)
    looks_like_iso_date = received_text.op("~")(r"^\d{4}-\d{2}-\d{2}")
    return case(
        (looks_like_iso_date, cast(received_text, DateTime(timezone=True))),
        else_=None,
    )


def _addresses(value: str) -> set[str]:
    return {address.strip().lower() for _, address in getaddresses([value or ""]) if address}


def _canonical_email_category(value: str) -> str:
    category = (value or "").strip().lower()
    return "phishing" if category == "malware" else category


def _email_matches_recipient(email_record: QuarantineEmail, identities: set[str]) -> bool:
    return bool(_addresses(email_record.recipient_list) & identities)


def _format_report_datetime(value) -> str:
    if not value:
        return ""
    if isinstance(value, datetime.datetime):
        if value.tzinfo is None:
            value = value.replace(tzinfo=datetime.timezone.utc)
        return value.astimezone(datetime.timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    return str(value)


def _gather_export_data(db: Session, req: ExportRequest, current_user: User) -> dict:
    dt_from = _parse_date(req.date_from)
    dt_to = _parse_date(req.date_to)
    if dt_from and dt_to and dt_from > dt_to:
        raise HTTPException(status_code=400, detail="date_from cannot be later than date_to")

    # Scope is based on the mailbox manager assignment used by the dashboard,
    # not organization-wide guesses or cached frontend values.
    admin_q = db.query(User).filter(User.role == UserRole.ADMIN.value)
    if current_user.role == UserRole.SUPERADMIN.value:
        if req.admin_ids is not None:
            admin_q = admin_q.filter(User.id.in_(req.admin_ids))
        admins = admin_q.order_by(User.username).all()
        mailbox_q = db.query(AdminMailbox)
        if req.admin_ids is not None:
            mailbox_q = mailbox_q.filter(AdminMailbox.assigned_to.in_([a.username for a in admins]))
    else:
        admins = [current_user]
        mailbox_q = db.query(AdminMailbox).filter(AdminMailbox.assigned_to == current_user.username)

    if req.mailbox_ids is not None:
        if not req.mailbox_ids:
            raise HTTPException(status_code=400, detail="Select at least one mailbox")
        mailbox_q = mailbox_q.filter(AdminMailbox.id.in_(req.mailbox_ids))

    mailboxes = mailbox_q.order_by(AdminMailbox.email).all()
    if req.mailbox_ids is not None:
        selected_ids = {mailbox.id for mailbox in mailboxes}
        missing_ids = set(req.mailbox_ids) - selected_ids
        if missing_ids:
            status_code = 403 if current_user.role != UserRole.SUPERADMIN.value else 404
            raise HTTPException(status_code=status_code, detail="One or more selected mailboxes are unavailable")
        manager_names = {mailbox.assigned_to for mailbox in mailboxes if mailbox.assigned_to}
        admins = [admin for admin in admins if admin.username in manager_names]
    mailbox_identities = {mailbox.email.strip().lower() for mailbox in mailboxes if mailbox.email}
    mailbox_by_email = {mailbox.email.strip().lower(): mailbox for mailbox in mailboxes if mailbox.email}

    # Only inbound, pipeline-classified messages belong in an anti-phishing
    # report. Drafts, sent mail and trash are separate webmail state.
    email_q = db.query(QuarantineEmail).filter(
        QuarantineEmail.label.notin_(["SENT", "DRAFT"]),
        or_(QuarantineEmail.status.is_(None), QuarantineEmail.status != "trash"),
    )
    received_at_expr = _report_received_at_expression(db)
    if dt_from:
        email_q = email_q.filter(received_at_expr >= dt_from)
    if dt_to:
        email_q = email_q.filter(received_at_expr < dt_to + datetime.timedelta(days=1))

    log_q = db.query(AuditLog)
    if dt_from:
        log_q = log_q.filter(AuditLog.created_at >= dt_from)
    if dt_to:
        log_q = log_q.filter(AuditLog.created_at < dt_to + datetime.timedelta(days=1))

    candidates = email_q.order_by(received_at_expr.desc().nullslast()).all()
    all_emails = [email for email in candidates if _email_matches_recipient(email, mailbox_identities)]
    total_email = len(all_emails)
    total_clean = sum(1 for e in all_emails if e.label == "CLEAN")
    total_warn = sum(1 for e in all_emails if e.label == "WARN")
    total_quarantine = sum(1 for e in all_emails if e.label == "QUARANTINE")
    total_phishing = sum(1 for e in all_emails if _canonical_email_category(e.category) == "phishing")
    total_spam = sum(1 for e in all_emails if (e.category or "").lower() == "spam")
    total_malware = 0
    avg_fused = sum(float(e.fused_score or 0) for e in all_emails) / total_email if total_email else 0
    avg_ml = sum(float(e.ml_probability or 0) for e in all_emails) / total_email if total_email else 0

    selected_org_ids = {admin.organization_id for admin in admins if admin.organization_id}
    user_q = db.query(User).filter(User.role == UserRole.USER.value)
    if current_user.role != UserRole.SUPERADMIN.value:
        if current_user.organization_id:
            user_q = user_q.filter(User.organization_id == current_user.organization_id)
        else:
            user_q = user_q.filter(func.lower(User.email).in_(mailbox_identities)) if mailbox_identities else user_q.filter(False)
    elif req.mailbox_ids is not None:
        user_q = user_q.filter(func.lower(User.email).in_(mailbox_identities)) if mailbox_identities else user_q.filter(False)
    elif req.admin_ids is not None:
        user_q = user_q.filter(User.organization_id.in_(selected_org_ids)) if selected_org_ids else user_q.filter(False)
    scoped_users = user_q.order_by(User.username).all()

    summary = {
        "total_admins": len(admins),
        "total_users": len(scoped_users),
        "total_organizations": len(selected_org_ids) if (req.admin_ids is not None or current_user.role != UserRole.SUPERADMIN.value) else db.query(Organization).count(),
        "total_mailboxes": len(mailboxes),
        "active_mailboxes": sum(1 for mailbox in mailboxes if mailbox.is_active),
        "total_emails": total_email,
        "total_clean": total_clean,
        "total_warn": total_warn,
        "total_quarantine": total_quarantine,
        "total_phishing": total_phishing,
        "total_spam": total_spam,
        "total_malware": total_malware,
        "avg_fused_score": round(avg_fused, 4),
        "avg_ml_probability": round(avg_ml, 4),
        "date_from": req.date_from or "All time",
        "date_to": req.date_to or "Present",
        "generated_at": _format_report_datetime(datetime.datetime.now(datetime.timezone.utc)),
        "scope": (
            f"Selected mailboxes ({len(mailboxes)})"
            if req.mailbox_ids is not None
            else "All managed mailboxes"
            if current_user.role == UserRole.SUPERADMIN.value
            else f"Mailboxes managed by {current_user.username}"
        ),
    }

    admin_data = []
    mailbox_data = []
    reported_mailbox_ids = set()
    for admin in admins:
        org_name = ""
        if admin.organization_id:
            org = db.query(Organization).filter(Organization.id == admin.organization_id).first()
            org_name = org.name if org else ""

        admin_mailboxes = [mailbox for mailbox in mailboxes if mailbox.assigned_to == admin.username]
        admin_identities = {mailbox.email.lower() for mailbox in admin_mailboxes}
        admin_emails = [email for email in all_emails if _email_matches_recipient(email, admin_identities)]
        admin_users = [user for user in scoped_users if admin.organization_id and user.organization_id == admin.organization_id]

        admin_logs = log_q.filter(AuditLog.user == admin.username
            ).order_by(AuditLog.created_at.desc()).limit(20).all()
        recent_actions = [{"action":log.action,"details":log.details or "",
                           "created_at":log.created_at.strftime("%Y-%m-%d %H:%M") if log.created_at else ""}
                          for log in admin_logs]

        for mb in admin_mailboxes:
            reported_mailbox_ids.add(mb.id)
            mb_emails = [email for email in all_emails if mb.email.lower() in _addresses(email.recipient_list)]
            mailbox_data.append({
                "admin": admin.username,
                "organization": org_name,
                "mailbox_email": mb.email,
                "domain": mb.domain,
                "is_active": mb.is_active,
                "created_at": mb.created_at.strftime("%Y-%m-%d") if mb.created_at else "",
                "email_stats": {
                    "total": len(mb_emails),
                    "clean": sum(1 for e in mb_emails if e.label == "CLEAN"),
                    "warn": sum(1 for e in mb_emails if e.label == "WARN"),
                    "quarantine": sum(1 for e in mb_emails if e.label == "QUARANTINE"),
                    "phishing": sum(1 for e in mb_emails if _canonical_email_category(e.category) == "phishing"),
                    "spam": sum(1 for e in mb_emails if (e.category or "").lower() == "spam"),
                    "malware": 0,
                },
            })

        admin_data.append({
            "username": admin.username,
            "role": admin.role, "organization": org_name,
            "is_active": admin.is_active,
            "user_count": len(admin_users), "mailbox_count": len(admin_mailboxes),
            "email_stats": {
                "total": len(admin_emails),
                "clean": sum(1 for e in admin_emails if e.label == "CLEAN"),
                "warn": sum(1 for e in admin_emails if e.label == "WARN"),
                "quarantine": sum(1 for e in admin_emails if e.label == "QUARANTINE"),
            },
            "recent_actions": recent_actions,
            "organization_id": admin.organization_id,
        })

    # A superadmin report must also expose real unassigned/orphaned mailboxes
    # instead of silently making the mailbox total disagree with the details.
    for mb in (mailbox for mailbox in mailboxes if mailbox.id not in reported_mailbox_ids):
        mb_emails = [email for email in all_emails if mb.email.lower() in _addresses(email.recipient_list)]
        mailbox_data.append({
            "admin": mb.assigned_to or "",
            "organization": "",
            "mailbox_email": mb.email,
            "domain": mb.domain,
            "is_active": mb.is_active,
            "created_at": mb.created_at.strftime("%Y-%m-%d") if mb.created_at else "",
            "email_stats": {
                "total": len(mb_emails),
                "clean": sum(1 for e in mb_emails if e.label == "CLEAN"),
                "warn": sum(1 for e in mb_emails if e.label == "WARN"),
                "quarantine": sum(1 for e in mb_emails if e.label == "QUARANTINE"),
                "phishing": sum(1 for e in mb_emails if _canonical_email_category(e.category) == "phishing"),
                "spam": sum(1 for e in mb_emails if (e.category or "").lower() == "spam"),
                "malware": 0,
            },
        })

    user_data = []
    if req.include_users:
        admin_by_org = {admin["organization_id"]: admin for admin in admin_data if admin["organization_id"]}
        for user in scoped_users:
            user_identity = {(user.email or "").strip().lower()} - {""}
            user_emails = [email for email in all_emails if _email_matches_recipient(email, user_identity)]
            owner = admin_by_org.get(user.organization_id, {})
            user_logs = log_q.filter(AuditLog.user == user.username).order_by(AuditLog.created_at.desc()).limit(10).all()
            user_data.append({
                "admin": owner.get("username", ""), "organization": owner.get("organization", ""),
                "username": user.username, "email": user.email or "", "is_active": user.is_active,
                "email_stats": {
                    "total": len(user_emails),
                    "clean": sum(1 for e in user_emails if e.label == "CLEAN"),
                    "warn": sum(1 for e in user_emails if e.label == "WARN"),
                    "quarantine": sum(1 for e in user_emails if e.label == "QUARANTINE"),
                },
                "recent_actions": [{
                    "action": log.action, "details": log.details or "",
                    "created_at": log.created_at.strftime("%Y-%m-%d %H:%M") if log.created_at else "",
                } for log in user_logs],
            })

    email_data = []
    if req.include_emails:
        org_names = {org.id: org.name for org in db.query(Organization).all()}
        for e in all_emails:
            org_name = ""
            if e.organization_id:
                org_name = org_names.get(e.organization_id, "")
            recipients = _addresses(e.recipient_list)
            matched_mailboxes = [mailbox_by_email[address] for address in recipients if address in mailbox_by_email]
            mailbox_labels = sorted({mailbox.email.lower() for mailbox in matched_mailboxes})
            manager_labels = sorted({mailbox.assigned_to for mailbox in matched_mailboxes if mailbox.assigned_to})
            attachments = []
            try:
                if e.attachments_json: attachments = json.loads(e.attachments_json)
            except Exception: pass
            has_attach = len(attachments) > 0
            has_malware = any(
                att.get("filename","").lower().endswith((".exe",".zip",".rar",".js",".vbs",".scr",".bat",".msi"))
                for att in attachments
            ) if attachments else False
            reasons = []
            if e.spf_result and e.spf_result.lower() not in ("pass", "n/a"):
                reasons.append(f"SPF:{e.spf_result}")
            if e.dkim_result and e.dkim_result.lower() not in ("pass", "n/a", "signed"):
                reasons.append(f"DKIM:{e.dkim_result}")
            if e.dmarc_result and e.dmarc_result.lower() not in ("pass", "n/a"):
                reasons.append(f"DMARC:{e.dmarc_result}")
            if has_malware:
                reasons.append("Malware extension")
            if e.routing_reason:
                reasons.append(e.routing_reason)
            elif e.xai_summary:
                reasons.append(e.xai_summary)
            email_data.append({
                "email_id": e.email_id, "subject": e.subject or "",
                "sender": e.sender or "", "recipient": e.recipient_list or "",
                "label": e.label or "", "category": _canonical_email_category(e.category),
                "organization": org_name, "mailbox": ", ".join(mailbox_labels),
                "mailbox_addresses": mailbox_labels,
                "admin": ", ".join(manager_labels),
                "received_at": _format_report_datetime(e.received_at),
                "fused_score": e.fused_score or 0, "sa_score": e.sa_score or 0,
                "ml_probability": e.ml_probability or 0, "anomaly_score": e.anomaly_score or 0,
                "has_attachment": has_attach, "has_malware_extension": has_malware,
                "spf_result": e.spf_result or "", "dkim_result": e.dkim_result or "",
                "dmarc_result": e.dmarc_result or "", "reasons": reasons,
                "model_version": e.model_version or "",
                "label_display": {"CLEAN":"Clean","WARN":"Suspicious","QUARANTINE":"Blocked"}.get(e.label, e.label or "Unknown"),
            })

    return {"summary": summary, "admins": admin_data, "users": user_data, "emails": email_data, "mailboxes": mailbox_data}


def _generate_pdf_report_legacy(data: dict):
    if not REPORTLAB_AVAILABLE:
        raise HTTPException(status_code=500, detail="reportlab not installed")
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    el = []; styles = getSampleStyleSheet()
    s = data["summary"]

    ts = ParagraphStyle('T',parent=styles['Heading1'],fontSize=20,
                         textColor=colors.HexColor('#1a73e8'),spaceAfter=6,alignment=1)
    ss = ParagraphStyle('S',parent=styles['Normal'],fontSize=9,
                         textColor=colors.grey,alignment=1,spaceAfter=20)

    el.append(Paragraph("CogniMail — Comprehensive Report", ts))
    el.append(Paragraph(f"Generated: {datetime.datetime.now(datetime.timezone.utc).strftime('%Y-%m-%d %H:%M:%S')} UTC | Period: {s['date_from']} — {s['date_to']}", ss))

    # Summary table
    st = Table([["Metric","Value"],
                ["Total Admins",str(s["total_admins"])],
                ["Total Users",str(s["total_users"])],
                ["Total Organizations",str(s["total_organizations"])],
                ["Total Emails",str(s["total_emails"])],
                ["Clean (Safe)",str(s["total_clean"])],
                ["Suspicious (Warn)",str(s["total_warn"])],
                ["Blocked (Quarantine)",str(s["total_quarantine"])]],
               colWidths=[3*inch,2*inch])
    st.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1a73e8')),
        ('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,0),9),
        ('GRID',(0,0),(-1,-1),0.4,colors.grey),
        ('ALIGN',(1,0),(1,-1),'CENTER'),
        ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4),
    ]))
    el.append(st); el.append(Spacer(1,16))

    h2 = ParagraphStyle('H2',parent=styles['Heading2'],fontSize=13,
                         textColor=colors.HexColor('#333'),spaceBefore=18,spaceAfter=8)
    h3 = ParagraphStyle('H3',parent=styles['Heading3'],fontSize=10,
                         textColor=colors.HexColor('#555'),spaceBefore=10,spaceAfter=4)
    cs = ParagraphStyle('C',parent=styles['Normal'],fontSize=7,leading=9)

    # Admins section
    el.append(Paragraph("Admin Details", h2))
    for a in data["admins"]:
        el.append(Paragraph(f"{a['username']} ({a['role']}) — {a['organization'] or 'Global'}", h3))
        el.append(Paragraph(
            f"Email: {a['email'] or '-'} | Status: {'Active' if a['is_active'] else 'Inactive'} | "
            f"Users: {a['user_count']} | Mailboxes: {a['mailbox_count']}", cs))
        es = a["email_stats"]
        at = Table([["Total","Clean","Suspicious","Blocked"],
                     [str(es["total"]),str(es["clean"]),str(es["warn"]),str(es["quarantine"])]],
                    colWidths=[1.1*inch]*4)
        at.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1a73e8')),
            ('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
            ('FONTSIZE',(0,0),(-1,-1),8),
            ('GRID',(0,0),(-1,-1),0.4,colors.grey),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
        ]))
        el.append(at)
        if a["recent_actions"]:
            el.append(Spacer(1,4))
            txt = "Recent: "+"; ".join(f"{x['action']}({x['created_at']})" for x in a["recent_actions"][:5])
            el.append(Paragraph(txt, cs))
        el.append(Spacer(1,10))

    # Users section
    if data["users"]:
        el.append(Paragraph("User Details", h2))
        ur = [["Admin","Org","Username","Email","Total","Clean","Warn","Blocked"]]
        for u in data["users"]:
            es = u["email_stats"]
            ur.append([u["admin"],u["organization"],u["username"],u["email"],
                       str(es["total"]),str(es["clean"]),str(es["warn"]),str(es["quarantine"])])
        if len(ur)>1:
            ut = Table(ur, colWidths=[0.9*inch,0.9*inch,0.8*inch,1.1*inch,0.5*inch,0.5*inch,0.6*inch,0.6*inch])
            ut.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1a73e8')),
                ('TEXTCOLOR',(0,0),(-1,0),colors.white),
                ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
                ('FONTSIZE',(0,0),(-1,-1),6),
                ('GRID',(0,0),(-1,-1),0.3,colors.grey),
                ('ALIGN',(4,0),(-1,-1),'CENTER'),
                ('TOPPADDING',(0,0),(-1,-1),2),('BOTTOMPADDING',(0,0),(-1,-1),2),
            ]))
            el.append(ut)
        el.append(Spacer(1,12))

    # Mailboxes section
    if data.get("mailboxes"):
        el.append(Paragraph("Mailbox Details", h2))
        mr = [["Admin","Org","Mailbox","Active","Total","Clean","Warn","Blocked"]]
        for m in data["mailboxes"]:
            es = m["email_stats"]
            mr.append([m["admin"],m["organization"],m["mailbox_email"],
                       "Yes" if m["is_active"] else "No",
                       str(es["total"]),str(es["clean"]),str(es["warn"]),str(es["quarantine"])])
        if len(mr)>1:
            mt = Table(mr, colWidths=[0.8*inch,0.8*inch,1.2*inch,0.5*inch,0.5*inch,0.5*inch,0.6*inch,0.6*inch])
            mt.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1a73e8')),
                ('TEXTCOLOR',(0,0),(-1,0),colors.white),
                ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
                ('FONTSIZE',(0,0),(-1,-1),6),
                ('GRID',(0,0),(-1,-1),0.3,colors.grey),
                ('ALIGN',(3,0),(-1,-1),'CENTER'),
                ('TOPPADDING',(0,0),(-1,-1),2),('BOTTOMPADDING',(0,0),(-1,-1),2),
            ]))
            el.append(mt)
        el.append(Spacer(1,12))

    # Emails section
    if data["emails"]:
        el.append(Paragraph("Email Details", h2))
        el.append(Paragraph(f"Total {len(data['emails'])} emails.", cs))
        el.append(Spacer(1,6))
        er = [["ID","Subject","Sender","Label","Category","Received","Score","Reasons"]]
        shown = 0
        for e in data["emails"]:
            if shown>=100: break
            rsn = "; ".join(e["reasons"][:2]) if e["reasons"] else "-"
            er.append([e["email_id"][:12]+".." if len(e["email_id"])>12 else e["email_id"],
                       (e["subject"][:40]+"..") if len(e["subject"])>40 else (e["subject"] or "-"),
                       (e["sender"][:25]+"..") if len(e["sender"])>25 else (e["sender"] or "-"),
                       e["label_display"], e["category"] or "-",
                       e["received_at"] or "-", f"{e['fused_score']:.2f}",
                       rsn[:50]])
            shown+=1
        if len(er)>1:
            et = Table(er, colWidths=[0.7*inch,1.1*inch,0.9*inch,0.5*inch,0.5*inch,0.7*inch,0.5*inch,1.1*inch])
            et.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1a73e8')),
                ('TEXTCOLOR',(0,0),(-1,0),colors.white),
                ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
                ('FONTSIZE',(0,0),(-1,-1),6),
                ('GRID',(0,0),(-1,-1),0.3,colors.grey),
                ('TOPPADDING',(0,0),(-1,-1),2),('BOTTOMPADDING',(0,0),(-1,-1),2),
            ]))
            el.append(et)
            if len(data["emails"])>100:
                el.append(Paragraph(f"... and {len(data['emails'])-100} more emails", cs))

    doc.build(el); buf.seek(0)
    return buf.getvalue()


def _generate_pdf_report(data: dict):
    """Generate a readable security report using only the gathered database rows."""
    if not REPORTLAB_AVAILABLE:
        raise HTTPException(status_code=500, detail="reportlab not installed")

    buf = io.BytesIO()
    page_size = landscape(A4)
    doc = SimpleDocTemplate(
        buf,
        pagesize=page_size,
        rightMargin=24,
        leftMargin=24,
        topMargin=30,
        bottomMargin=30,
        title="CogniMail Security and Email Classification Report",
        author="CogniMail",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle", parent=styles["Heading1"], fontName="Helvetica-Bold",
        fontSize=18, leading=22, textColor=colors.HexColor("#111827"), spaceAfter=5,
    )
    subtitle_style = ParagraphStyle(
        "ReportSubtitle", parent=styles["Normal"], fontSize=8, leading=11,
        textColor=colors.HexColor("#4B5563"), spaceAfter=10,
    )
    heading_style = ParagraphStyle(
        "ReportHeading", parent=styles["Heading2"], fontName="Helvetica-Bold",
        fontSize=11, leading=14, textColor=colors.HexColor("#1D4ED8"),
        spaceBefore=10, spaceAfter=6,
    )
    cell_style = ParagraphStyle(
        "ReportCell", parent=styles["Normal"], fontSize=6.4, leading=8,
        textColor=colors.HexColor("#1F2937"), wordWrap="CJK",
    )
    header_style = ParagraphStyle(
        "ReportHeader", parent=cell_style, fontName="Helvetica-Bold",
        textColor=colors.white, alignment=1,
    )

    def paragraph(value, style=cell_style):
        text = "-" if value is None or value == "" else str(value)
        return Paragraph(escape(text).replace("\n", "<br/>"), style)

    def styled_table(rows, widths, numeric_from=None):
        table = Table(rows, colWidths=widths, repeatRows=1, hAlign="LEFT")
        commands = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]
        for row_index in range(1, len(rows)):
            if row_index % 2 == 0:
                commands.append(("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#F8FAFC")))
        if numeric_from is not None:
            commands.append(("ALIGN", (numeric_from, 1), (-1, -1), "CENTER"))
        table.setStyle(TableStyle(commands))
        return table

    summary = data["summary"]
    elements = [
        Paragraph("CogniMail Security &amp; Email Classification Report", title_style),
        Paragraph(
            f"Generated {escape(summary['generated_at'])} | Period: "
            f"{escape(str(summary['date_from']))} to {escape(str(summary['date_to']))} | "
            f"Scope: {escape(summary['scope'])}",
            subtitle_style,
        ),
        Paragraph(
            "Source: live CogniMail database. This report includes inbound messages processed by the "
            "anti-phishing and spam pipeline; sent mail, drafts, trash, and sample rows are excluded.",
            subtitle_style,
        ),
        Paragraph("Executive Summary", heading_style),
    ]

    metrics = [
        ("Processed email", summary["total_emails"]),
        ("Managed mailbox", f"{summary['total_mailboxes']} ({summary['active_mailboxes']} active)"),
        ("Clean", summary["total_clean"]),
        ("Suspicious", summary["total_warn"]),
        ("Blocked", summary["total_quarantine"]),
        ("Phishing", summary["total_phishing"]),
        ("Spam", summary["total_spam"]),
        ("Average fused score", f"{summary['avg_fused_score']:.4f}"),
        ("Average ML probability", f"{summary['avg_ml_probability']:.4f}"),
    ]
    metric_rows = [[paragraph("Metric", header_style), paragraph("Value", header_style),
                    paragraph("Metric", header_style), paragraph("Value", header_style)]]
    for index in range(0, len(metrics), 2):
        left = metrics[index]
        right = metrics[index + 1] if index + 1 < len(metrics) else ("", "")
        metric_rows.append([paragraph(left[0]), paragraph(left[1]), paragraph(right[0]), paragraph(right[1])])
    elements.extend([styled_table(metric_rows, [2.25*inch, 1.05*inch, 2.25*inch, 1.05*inch]), Spacer(1, 8)])

    if data.get("mailboxes"):
        elements.append(Paragraph("Mailbox Coverage", heading_style))
        mailbox_rows = [[paragraph(x, header_style) for x in
                         ["Mailbox", "Admin", "Status", "Processed", "Clean", "Suspicious", "Blocked", "Phishing", "Spam"]]]
        for mailbox in data["mailboxes"]:
            stats = mailbox["email_stats"]
            mailbox_rows.append([
                paragraph(mailbox["mailbox_email"]), paragraph(mailbox["admin"] or "Unassigned"),
                paragraph("Active" if mailbox["is_active"] else "Inactive"), paragraph(stats["total"]),
                paragraph(stats["clean"]), paragraph(stats["warn"]), paragraph(stats["quarantine"]),
                paragraph(stats["phishing"]), paragraph(stats["spam"]),
            ])
        elements.extend([styled_table(mailbox_rows, [1.8*inch, .9*inch, .7*inch, .65*inch,
                         .55*inch, .72*inch, .6*inch, .65*inch, .55*inch], 3), Spacer(1, 8)])

    if data.get("users"):
        elements.append(Paragraph("Mailbox User Accounts", heading_style))
        user_rows = [[paragraph(x, header_style) for x in
                      ["Admin", "Username", "Email", "Status", "Processed", "Clean", "Suspicious", "Blocked"]]]
        for user in data["users"]:
            stats = user["email_stats"]
            user_rows.append([
                paragraph(user["admin"]), paragraph(user["username"]), paragraph(user["email"]),
                paragraph("Active" if user["is_active"] else "Inactive"), paragraph(stats["total"]),
                paragraph(stats["clean"]), paragraph(stats["warn"]), paragraph(stats["quarantine"]),
            ])
        elements.extend([styled_table(user_rows, [1.0*inch, 1.0*inch, 2.2*inch, .7*inch,
                         .7*inch, .55*inch, .75*inch, .6*inch], 4), Spacer(1, 8)])

    # Keep every mailbox in its own report section. This avoids mixing results
    # when a superadmin exports several customer addresses at once.
    for mailbox_index, mailbox in enumerate(data.get("mailboxes", [])):
        if mailbox_index:
            elements.append(PageBreak())
        address = mailbox["mailbox_email"].lower()
        mailbox_emails = [
            row for row in data.get("emails", [])
            if address in row.get("mailbox_addresses", [])
        ]
        stats = mailbox["email_stats"]
        elements.append(Paragraph(f"Mailbox Report: {escape(address)}", heading_style))
        elements.append(Paragraph(
            f"Responsible admin: {escape(mailbox['admin'] or 'Unassigned')} | "
            f"Status: {'Active' if mailbox['is_active'] else 'Inactive'} | "
            f"Processed: {stats['total']} | Clean: {stats['clean']} | "
            f"Blocked: {stats['quarantine']} | Phishing: {stats['phishing']} | "
            f"Spam: {stats['spam']}",
            subtitle_style,
        ))
        if not mailbox_emails:
            elements.append(Paragraph(
                "No inbound records matched this mailbox and selected period.", subtitle_style
            ))
            continue
        email_rows = [[paragraph(x, header_style) for x in
                       ["Received (UTC)", "Subject", "Sender", "Result / Category",
                        "Scores F / ML / SA", "Authentication", "Detection evidence"]]]
        for email_row in mailbox_emails:
            evidence = "; ".join(email_row["reasons"]) or "No additional rule evidence"
            email_rows.append([
                paragraph(email_row["received_at"]),
                paragraph(email_row["subject"] or "(No subject)"),
                paragraph(email_row["sender"]),
                paragraph(f"{email_row['label_display']} / {email_row['category'] or 'uncategorized'}"),
                paragraph(f"{float(email_row['fused_score']):.3f} / {float(email_row['ml_probability']):.3f} / {float(email_row['sa_score']):.2f}"),
                paragraph(f"SPF {email_row['spf_result'] or 'N/A'}; DKIM {email_row['dkim_result'] or 'N/A'}; DMARC {email_row['dmarc_result'] or 'N/A'}"),
                paragraph(evidence),
            ])
        elements.append(styled_table(email_rows, [1.1*inch, 1.9*inch, 1.55*inch,
                        1.05*inch, .95*inch, 1.45*inch, 2.2*inch]))

    if not data.get("mailboxes"):
        elements.extend([Paragraph("Mailbox Reports", heading_style),
                         Paragraph("No managed mailbox matched the selected scope.", subtitle_style)])

    def add_page_number(canvas, report_doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#64748B"))
        canvas.drawString(24, 15, "CogniMail | ML-Powered Anti-Phishing and Spam Filtering")
        canvas.drawRightString(page_size[0] - 24, 15, f"Page {report_doc.page}")
        canvas.restoreState()

    doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)
    buf.seek(0)
    return buf.getvalue()


def _generate_excel_report(data: dict):
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    wb = openpyxl.Workbook()
    hf = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
    hfn = Font(bold=True, color="FFFFFF", size=10)
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    def sh(ws, n):
        for c in range(1,n+1):
            cell = ws.cell(row=1,column=c)
            cell.fill = hf; cell.font = hfn; cell.alignment = ha
    def aw(ws):
        for cc in ws.columns:
            letter = cc[0].column_letter
            mx = min(max((len(str(c.value)) for c in cc if c.value), default=0)+3, 50)
            ws.column_dimensions[letter].width = mx

    s = data["summary"]
    ws1 = wb.active; ws1.title = "Summary"
    for row in [("Metric","Value"),
                ("Total Admins",s["total_admins"]),
                ("Total Users",s["total_users"]),
                ("Total Organizations",s["total_organizations"]),
                ("Managed Mailboxes",s["total_mailboxes"]),
                ("Active Mailboxes",s["active_mailboxes"]),
                ("Total Emails",s["total_emails"]),
                ("Clean (Safe)",s["total_clean"]),
                ("Suspicious (Warn)",s["total_warn"]),
                ("Blocked (Quarantine)",s["total_quarantine"]),
                ("Phishing",s["total_phishing"]),
                ("Spam",s["total_spam"]),
                ("Average Fused Score",s["avg_fused_score"]),
                ("Average ML Probability",s["avg_ml_probability"]),
                ("Period From",s["date_from"]),
                ("Period To",s["date_to"]),
                ("Scope",s["scope"]),
                ("Generated",s["generated_at"])]:
        ws1.append(list(row))
    sh(ws1,2); aw(ws1)

    ws2 = wb.create_sheet("Admins")
    h2 = ["Username","Role","Active","Users","Mailboxes",
          "Total Emails","Clean","Suspicious","Blocked","Recent Activity"]
    ws2.append(h2); sh(ws2,len(h2))
    for a in data["admins"]:
        recent = "; ".join(f"{x['action']}({x['created_at']})" for x in a["recent_actions"][:5])
        es = a["email_stats"]
        ws2.append([a["username"],a["role"],
                    "Yes" if a["is_active"] else "No",
                    a["user_count"],a["mailbox_count"],
                    es["total"],es["clean"],es["warn"],es["quarantine"],recent])
    aw(ws2)

    if data["users"]:
        ws3 = wb.create_sheet("Users")
        h3 = ["Admin","Username","Email","Active",
              "Total Emails","Clean","Suspicious","Blocked","Recent Activity"]
        ws3.append(h3); sh(ws3,len(h3))
        for u in data["users"]:
            recent = "; ".join(f"{x['action']}({x['created_at']})" for x in u["recent_actions"][:3])
            es = u["email_stats"]
            ws3.append([u["admin"],u["username"],u["email"],
                        "Yes" if u["is_active"] else "No",
                        es["total"],es["clean"],es["warn"],es["quarantine"],recent])
        aw(ws3)

    if data.get("mailboxes"):
        wsm = wb.create_sheet("Mailboxes")
        hm = ["Admin","Mailbox","Domain","Active","Created",
              "Total Emails","Clean","Suspicious","Blocked","Phishing","Spam"]
        wsm.append(hm); sh(wsm,len(hm))
        for m in data["mailboxes"]:
            es = m["email_stats"]
            wsm.append([m["admin"],m["mailbox_email"],m["domain"],
                        "Yes" if m["is_active"] else "No",m["created_at"],
                        es["total"],es["clean"],es["warn"],es["quarantine"],
                        es["phishing"],es["spam"]])
        aw(wsm)

    used_sheet_names = set(wb.sheetnames)
    for mailbox_index, mailbox in enumerate(data.get("mailboxes", []), start=1):
        address = mailbox["mailbox_email"].lower()
        base_name = "Mail " + "".join(
            character if character not in '[]:*?/\\' else "_" for character in address
        )
        base_name = base_name[:31] or f"Mailbox {mailbox_index}"
        sheet_name = base_name
        suffix = 2
        while sheet_name in used_sheet_names:
            tail = f" {suffix}"
            sheet_name = base_name[:31-len(tail)] + tail
            suffix += 1
        used_sheet_names.add(sheet_name)
        ws4 = wb.create_sheet(sheet_name)
        ws4.append(["Mailbox", address])
        ws4.append(["Responsible Admin", mailbox["admin"] or "Unassigned"])
        ws4.append(["Status", "Active" if mailbox["is_active"] else "Inactive"])
        ws4.append([])
        h4 = ["Email ID","Subject","Sender","Recipient","Label","Category","Received At",
              "Fused Score","SA Score","ML Prob","Anomaly","Has Attachment",
              "Malware Extension","SPF","DKIM","DMARC","Model Version","Reasons"]
        ws4.append(h4)
        for cell in ws4[5]:
            cell.fill = hf; cell.font = hfn; cell.alignment = ha
        mailbox_emails = [
            email for email in data.get("emails", [])
            if address in email.get("mailbox_addresses", [])
        ]
        for e in mailbox_emails:
            ws4.append([e["email_id"],e["subject"],e["sender"],e["recipient"],
                        e["label_display"],e["category"],e["received_at"],
                        e["fused_score"],e["sa_score"],e["ml_probability"],e["anomaly_score"],
                        "Yes" if e["has_attachment"] else "No",
                        "Yes" if e["has_malware_extension"] else "No",
                        e["spf_result"],e["dkim_result"],e["dmarc_result"],e["model_version"],
                        "; ".join(e["reasons"])])
        aw(ws4)

    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out.getvalue()


@router.post("/export/generate")
def generate_export(
    req: ExportRequest,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Generate a live-data comprehensive report with role and mailbox filters."""
    data = _gather_export_data(db, req, current_user)
    fmt = req.format.lower()
    ts = datetime.datetime.now(datetime.UTC).strftime("%Y%m%d_%H%M%S")
    if fmt == "pdf":
        content = _generate_pdf_report(data)
        return StreamingResponse(iter([content]), media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=cognimail_report_{ts}.pdf"})
    elif fmt == "excel":
        content = _generate_excel_report(data)
        return StreamingResponse(iter([content]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=cognimail_report_{ts}.xlsx"})
    else:
        raise HTTPException(status_code=400, detail=f"Unsupported format: {fmt}. Use pdf or excel.")


@router.get("/admins/list")
def list_admins_for_export(
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Return list of admins for export filter selection.
    Superadmin sees all admins; admin sees only those in their org (self).
    """
    q = db.query(User).filter(User.role == UserRole.ADMIN.value)
    if current_user.role != UserRole.SUPERADMIN.value and current_user.organization_id:
        q = q.filter(User.organization_id == current_user.organization_id)
    admins = q.order_by(User.username).all()
    return [{"id":a.id,"username":a.username,
             "role":a.role,"organization_id":a.organization_id} for a in admins]


@router.get("/search")
async def api_global_search(
    q: str = Query("", min_length=1, max_length=100),
    _admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Unified global search across users, emails, and audit logs."""
    query = q.strip()
    if not query:
        return {"pages": [], "users": [], "emails": [], "logs": []}

    like = f"%{query}%"

    # Admin search results are tenant-scoped.
    restrict_org_id = None
    if _admin.role == UserRole.ADMIN.value and _admin.organization_id:
        restrict_org_id = _admin.organization_id

    # Users
    user_q = db.query(User).filter(
        or_(User.username.ilike(like), User.email.ilike(like))
    )
    if restrict_org_id:
        user_q = user_q.filter(User.organization_id == restrict_org_id)
    user_rows = user_q.limit(5).all()
    users = [
        {"username": u.username, "email": u.email or "", "role": u.role}
        for u in user_rows
    ]

    # Emails (subject, sender) — scoped to org
    email_q = db.query(QuarantineEmail).filter(
        or_(QuarantineEmail.subject.ilike(like), QuarantineEmail.sender.ilike(like))
    )
    if restrict_org_id:
        email_q = email_q.filter(QuarantineEmail.organization_id == restrict_org_id)
    email_rows = email_q.order_by(QuarantineEmail.received_at.desc().nullslast()).limit(5).all()
    emails = [
        {
            "email_id": e.email_id,
            "subject": e.subject or "",
            "sender": e.sender or "",
            "label": e.label or "",
            "received_at": e.received_at or "",
        }
        for e in email_rows
    ]

    # Audit logs — scoped to org users
    log_q = db.query(AuditLog).filter(
        or_(AuditLog.user.ilike(like), AuditLog.action.ilike(like), AuditLog.details.ilike(like))
    )
    if restrict_org_id:
        org_usernames = [u.username for u in user_rows] or ["__no_match__"]
        log_q = log_q.filter(AuditLog.user.in_(org_usernames))
    log_rows = log_q.order_by(AuditLog.created_at.desc().nullslast()).limit(5).all()
    logs = [
        {
            "id": l.id,
            "user": l.user or "",
            "action": l.action or "",
            "details": l.details or "",
            "created_at": l.created_at.isoformat() if l.created_at else "",
        }
        for l in log_rows
    ]

    return {"users": users, "emails": emails, "logs": logs}
