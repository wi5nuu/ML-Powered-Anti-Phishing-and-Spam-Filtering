import asyncio
import io
import os
import re
import tempfile
import unittest
import zipfile
from pathlib import Path
from unittest.mock import AsyncMock, patch

os.environ["ENV"] = "testing"
os.environ["DASHBOARD_DB_URL"] = "sqlite:///:memory:"
os.environ.setdefault("DASHBOARD_SECRET_KEY", "test-secret-key-that-is-long-enough")

from fastapi.testclient import TestClient  # noqa: E402
from PIL import Image  # noqa: E402
from openpyxl import load_workbook  # noqa: E402
from pypdf import PdfReader  # noqa: E402
from sqlalchemy import create_engine  # noqa: E402
from sqlalchemy.orm import sessionmaker  # noqa: E402
from sqlalchemy.pool import StaticPool  # noqa: E402

from dashboard.app import ConnectionManager, app, build_detection_explanation  # noqa: E402
from dashboard.auth import create_access_token, hash_password  # noqa: E402
from dashboard.database import get_db  # noqa: E402
from database.models import (  # noqa: E402
    AdminMailbox,
    AdminMailboxAccess,
    AuditLog,
    AuditTrail,
    Base,
    Feedback,
    Organization,
    QuarantineEmail,
    Report,
    TrainingSample,
    User,
)


test_engine = create_engine(
    "sqlite://",
    connect_args={"check_same_thread": False},
    poolclass=StaticPool,
)
TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=test_engine)


def override_get_db():
    db = TestingSessionLocal()
    try:
        yield db
    finally:
        db.close()


class DashboardUserFlowTests(unittest.TestCase):
    def setUp(self):
        Base.metadata.drop_all(test_engine)
        Base.metadata.create_all(test_engine)
        app.dependency_overrides[get_db] = override_get_db
        self.db = TestingSessionLocal()
        self.organization = Organization(name="Route Test Organization")
        self.superadmin = User(
            username="route-test-superadmin",
            email="route-superadmin@example.test",
            hashed_password=hash_password("test-password-123"),
            role="superadmin",
            is_active=True,
        )
        self.db.add_all([self.organization, self.superadmin])
        self.db.commit()
        self.client = TestClient(app, base_url="http://localhost")
        # Authenticate directly so the suite itself does not exhaust the
        # production login rate limit while creating a fresh client per test.
        token = create_access_token({"sub": self.superadmin.username, "role": self.superadmin.role})
        self.client.cookies.set("access_token", token)

    def tearDown(self):
        self.client.close()
        self.db.close()
        app.dependency_overrides.pop(get_db, None)
        Base.metadata.drop_all(test_engine)

    def test_api_routes_have_no_duplicate_method_and_shape(self):
        seen = {}
        duplicates = []
        for route in app.routes:
            normalized = re.sub(r"\{[^}]+\}", "{}", getattr(route, "path", ""))
            for method in getattr(route, "methods", set()) or set():
                if method in {"HEAD", "OPTIONS"}:
                    continue
                key = (method, normalized)
                if key in seen:
                    duplicates.append((key, seen[key], route.name))
                seen[key] = route.name
        self.assertEqual(duplicates, [])

    def test_new_email_websocket_refreshes_matching_mailbox_for_every_category(self):
        notification_manager = ConnectionManager()
        matching_mailbox = AsyncMock()
        other_mailbox = AsyncMock()
        superadmin_socket = AsyncMock()
        notification_manager.active_connections = {
            matching_mailbox: {
                "role": "mailbox",
                "mailbox_email": "bantuan@zenime.my.id",
            },
            other_mailbox: {
                "role": "mailbox",
                "mailbox_email": "other@zenime.my.id",
            },
            superadmin_socket: {
                "role": "superadmin",
                "mailbox_email": "",
            },
        }

        clean_event = {
            "type": "email_processed",
            "email_id": "clean-notification-test",
            "label": "CLEAN",
            "category": "clean",
            "status": "released",
            "recipients": ["bantuan@zenime.my.id"],
        }
        asyncio.run(notification_manager.broadcast(clean_event))
        matching_mailbox.send_json.assert_awaited_once_with(clean_event)
        other_mailbox.send_json.assert_not_awaited()
        superadmin_socket.send_json.assert_not_awaited()

        matching_mailbox.send_json.reset_mock()
        for processed_event in (
            {**clean_event, "email_id": "spam-test", "label": "QUARANTINE", "category": "spam", "status": "pending"},
            {**clean_event, "email_id": "phishing-test", "label": "QUARANTINE", "category": "phishing", "status": "pending"},
        ):
            asyncio.run(notification_manager.broadcast(processed_event))
            matching_mailbox.send_json.assert_awaited_once_with(processed_event)
            matching_mailbox.send_json.reset_mock()

        asyncio.run(notification_manager.broadcast({"type": "ping"}))
        matching_mailbox.send_json.assert_not_awaited()

    def test_bulk_delete_refills_first_page_with_remaining_emails(self):
        messages = [
            QuarantineEmail(
                email_id=f"bulk-delete-{index:02d}",
                subject=f"Message {index:02d}",
                label="CLEAN",
                category="clean",
                status="released",
                fused_score=0.1,
                sender="sender@example.test",
                recipient_list="recipient@example.test",
            )
            for index in range(70)
        ]
        self.db.add_all(messages)
        self.db.commit()

        first_page = self.client.get("/api/emails", params={"page": 1, "page_size": 50})
        self.assertEqual(first_page.status_code, 200, first_page.text)
        self.assertEqual(first_page.json()["total"], 70)
        selected_ids = [row["email_id"] for row in first_page.json()["emails"]]

        deleted = self.client.post(
            "/api/emails/bulk-delete",
            json={"email_ids": selected_ids},
        )
        self.assertEqual(deleted.status_code, 200, deleted.text)
        self.assertEqual(deleted.json()["processed"], 50)
        self.assertEqual(deleted.json()["moved_to_trash"], 50)

        refilled = self.client.get("/api/emails", params={"page": 1, "page_size": 50})
        self.assertEqual(refilled.status_code, 200, refilled.text)
        self.assertIn("no-store", refilled.headers.get("cache-control", ""))
        self.assertEqual(refilled.json()["total"], 20)
        self.assertEqual(len(refilled.json()["emails"]), 20)

    def test_xai_never_turns_unverified_authentication_into_failure(self):
        email = QuarantineEmail(
            email_id="truthful-xai-message",
            label="QUARANTINE",
            category="phishing",
            status="pending",
            fused_score=0.91,
            ml_probability=0.9824,
            anomaly_score=0.7378,
            spf_result="N/A",
            dkim_result="N/A",
            dmarc_result="N/A",
            xai_summary=(
                "SpamProb=0.98; Urgency-Score:0.75; SPF:FAIL; DKIM:FAIL; "
                "Top-SHAP:password"
            ),
            shap_json=(
                '{"source":"xgboost_pred_contribs","prediction_probability":0.9824,'
                '"is_anomaly":true,"features":[{"name":"password",'
                '"shap":1.2345,"direction":"spam"}]}'
            ),
        )
        reasons, human_reasons, shap_data = build_detection_explanation(email)
        self.assertFalse(any(reason["key"] in {"SPF", "DKIM"} for reason in reasons))
        self.assertFalse(any("gagal" in reason.lower() or "tidak valid" in reason.lower() for reason in human_reasons))
        self.assertIn("Model supervised menghasilkan probabilitas spam 98.24%.", human_reasons)
        self.assertEqual(shap_data["features"][0]["name"], "password")

        email.spf_result = "FAIL"
        reasons, human_reasons, _ = build_detection_explanation(email)
        self.assertIn({"key": "SPF", "value": "FAIL"}, reasons)
        self.assertTrue(any("SPF" in reason and "gagal" in reason for reason in human_reasons))

    def test_main_login_accepts_admin_created_mailbox_credentials(self):
        admin = User(
            username="mailbox-owner",
            email="owner@example.test",
            hashed_password=hash_password("Admin-password-123"),
            role="admin",
            is_active=True,
        )
        mailbox = AdminMailbox(
            email="support@example.test",
            domain="example.test",
            password_hash=hash_password("Mailbox-password-123"),
            sender_name="Support",
            assigned_to=admin.username,
            created_by=admin.username,
            is_active=True,
        )
        self.db.add_all([admin, mailbox])
        self.db.commit()

        mailbox_client = TestClient(app, base_url="http://localhost")
        try:
            login = mailbox_client.post(
                "/api/auth/login",
                data={"username": "SUPPORT@EXAMPLE.TEST", "password": "Mailbox-password-123"},
            )
            self.assertEqual(login.status_code, 200, login.text)
            self.assertEqual(login.json()["role"], "mailbox")
            self.assertEqual(login.json()["mailbox_email"], "support@example.test")

            me = mailbox_client.get("/api/auth/me")
            self.assertEqual(me.status_code, 200, me.text)
            self.assertTrue(me.json()["authenticated"])
            self.assertEqual(me.json()["user"]["role"], "mailbox")
        finally:
            mailbox_client.close()

    def test_mailbox_user_cannot_see_threats_until_admin_releases_them(self):
        mailbox = AdminMailbox(
            email="safe-only@example.test",
            domain="example.test",
            password_hash=hash_password("Mailbox-password-123"),
            sender_name="Safe Only",
            created_by=self.superadmin.username,
            is_active=True,
        )
        clean = QuarantineEmail(
            email_id="safe-visible-message",
            subject="Visible clean message",
            label="CLEAN",
            category="clean",
            status="released",
            fused_score=0.0,
            sender="sender@example.test",
            recipient_list=mailbox.email,
        )
        threat = QuarantineEmail(
            email_id="hidden-phishing-message",
            subject="Secret phishing marker",
            label="QUARANTINE",
            category="phishing",
            status="quarantined",
            fused_score=0.95,
            sender="attacker@example.test",
            recipient_list=mailbox.email,
        )
        inconsistent_quarantine = QuarantineEmail(
            email_id="hidden-inconsistent-quarantine",
            subject="Quarantine with stale clean label",
            label="CLEAN",
            category="spam",
            status="quarantined",
            fused_score=0.91,
            sender="stale-label-attacker@example.test",
            recipient_list=mailbox.email,
        )
        delivered_warning = QuarantineEmail(
            email_id="visible-warning-message",
            subject="Review this warning",
            label="WARN",
            category="warn",
            status="warn_delivered",
            fused_score=0.72,
            ml_probability=0.81,
            shap_json=(
                '{"source":"xgboost_pred_contribs",'
                '"prediction_probability":0.81,'
                '"features":[],"warning_header":"CogniMail WARN; Fused=0.7200; ML=0.8100"}'
            ),
            sender="uncertain@example.test",
            recipient_list=mailbox.email,
        )
        pending_warning = QuarantineEmail(
            email_id="hidden-pending-warning",
            subject="Pending warning must stay hidden",
            label="WARN",
            category="warn",
            status="pending",
            fused_score=0.70,
            sender="pending@example.test",
            recipient_list=mailbox.email,
        )
        self.db.add_all([
            mailbox, clean, threat, inconsistent_quarantine,
            delivered_warning, pending_warning,
        ])
        self.db.commit()

        mailbox_client = TestClient(app, base_url="http://localhost")
        try:
            login = mailbox_client.post(
                "/api/auth/login",
                data={"username": mailbox.email, "password": "Mailbox-password-123"},
            )
            self.assertEqual(login.status_code, 200, login.text)

            inbox = mailbox_client.get("/api/emails", params={"q": "Secret phishing marker"})
            self.assertEqual(inbox.status_code, 200, inbox.text)
            self.assertEqual(inbox.json()["emails"], [])
            all_mail = mailbox_client.get(
                "/api/emails",
                params={"folder": "all", "q": "Quarantine with stale clean label"},
            )
            self.assertEqual(all_mail.status_code, 200, all_mail.text)
            self.assertEqual(all_mail.json()["emails"], [])
            category = mailbox_client.get("/api/emails", params={"category": "phishing"})
            self.assertEqual(category.status_code, 403, category.text)
            warning_category = mailbox_client.get("/api/emails", params={"category": "warn"})
            self.assertEqual(warning_category.status_code, 200, warning_category.text)
            self.assertEqual(
                [row["email_id"] for row in warning_category.json()["emails"]],
                [delivered_warning.email_id],
            )
            warning_detail = mailbox_client.get(f"/api/emails/{delivered_warning.email_id}")
            self.assertEqual(warning_detail.status_code, 200, warning_detail.text)
            self.assertEqual(
                warning_detail.json()["warning_xai_header"],
                "CogniMail WARN; Fused=0.7200; ML=0.8100",
            )
            pending_warning_detail = mailbox_client.get(f"/api/emails/{pending_warning.email_id}")
            self.assertEqual(pending_warning_detail.status_code, 404, pending_warning_detail.text)
            detail = mailbox_client.get(f"/api/emails/{threat.email_id}")
            self.assertEqual(detail.status_code, 404, detail.text)
            inconsistent_detail = mailbox_client.get(
                f"/api/emails/{inconsistent_quarantine.email_id}"
            )
            self.assertEqual(inconsistent_detail.status_code, 404, inconsistent_detail.text)
            stats = mailbox_client.get("/api/stats")
            self.assertEqual(stats.status_code, 200, stats.text)
            self.assertEqual(stats.json()["total"], 2)
            self.assertEqual(stats.json()["quarantine"], 0)
            self.assertEqual(stats.json()["warn"], 1)
            self.assertEqual(stats.json()["categories"]["warn"], 1)
            self.assertNotIn("phishing", stats.json()["categories"])

            admin_list = self.client.get(
                "/api/emails",
                params={"mailbox_id": mailbox.id, "category": "phishing"},
            )
            self.assertEqual(admin_list.status_code, 200, admin_list.text)
            self.assertEqual([row["email_id"] for row in admin_list.json()["emails"]], [threat.email_id])
            release = self.client.post(f"/api/emails/{threat.email_id}/release")
            self.assertEqual(release.status_code, 200, release.text)

            released = mailbox_client.get(f"/api/emails/{threat.email_id}")
            self.assertEqual(released.status_code, 200, released.text)
            self.assertEqual(released.json()["label"], "CLEAN")
            self.assertEqual(released.json()["reasons"], [])
        finally:
            mailbox_client.close()

    def test_warn_has_dedicated_review_filter_and_is_not_counted_as_spam(self):
        mailbox = AdminMailbox(
            email="warning-review@example.test",
            domain="example.test",
            created_by=self.superadmin.username,
            is_active=True,
        )
        warning = QuarantineEmail(
            email_id="dedicated-warning-message",
            subject="Needs warning review",
            label="WARN",
            category="phishing",
            status="pending",
            fused_score=0.72,
            sender="uncertain@example.test",
            recipient_list=mailbox.email,
        )
        spam = QuarantineEmail(
            email_id="confirmed-spam-message",
            subject="Confirmed spam",
            label="QUARANTINE",
            category="spam",
            status="pending",
            fused_score=0.96,
            sender="spam@example.test",
            recipient_list=mailbox.email,
        )
        legacy_malware = QuarantineEmail(
            email_id="legacy-malware-now-phishing",
            subject="Legacy attachment threat",
            label="QUARANTINE",
            category="malware",
            status="pending",
            fused_score=0.99,
            sender="payload@example.test",
            recipient_list=mailbox.email,
        )
        self.db.add_all([mailbox, warning, spam, legacy_malware])
        self.db.commit()

        warning_list = self.client.get(
            "/api/emails", params={"mailbox_id": mailbox.id, "category": "warn"}
        )
        self.assertEqual(warning_list.status_code, 200, warning_list.text)
        self.assertEqual(
            [row["email_id"] for row in warning_list.json()["emails"]],
            [warning.email_id],
        )

        spam_list = self.client.get(
            "/api/emails", params={"mailbox_id": mailbox.id, "category": "spam"}
        )
        self.assertEqual(spam_list.status_code, 200, spam_list.text)
        self.assertEqual(
            [row["email_id"] for row in spam_list.json()["emails"]],
            [spam.email_id],
        )

        phishing_list = self.client.get(
            "/api/emails", params={"mailbox_id": mailbox.id, "category": "phishing"}
        )
        self.assertEqual(phishing_list.status_code, 200, phishing_list.text)
        self.assertEqual(
            [row["email_id"] for row in phishing_list.json()["emails"]],
            [legacy_malware.email_id],
        )
        self.assertEqual(phishing_list.json()["emails"][0]["category"], "phishing")

        legacy_alias = self.client.get(
            "/api/emails", params={"mailbox_id": mailbox.id, "category": "malware"}
        )
        self.assertEqual(legacy_alias.status_code, 200, legacy_alias.text)
        self.assertEqual(
            [row["email_id"] for row in legacy_alias.json()["emails"]],
            [legacy_malware.email_id],
        )
        self.assertEqual(legacy_alias.json()["emails"][0]["category"], "phishing")

        quarantine_alias = self.client.get(
            "/api/admin/quarantine", params={"mailbox": mailbox.email, "category": "malware"}
        )
        self.assertEqual(quarantine_alias.status_code, 200, quarantine_alias.text)
        self.assertEqual(
            [row["email_id"] for row in quarantine_alias.json()["emails"]],
            [legacy_malware.email_id],
        )
        self.assertEqual(quarantine_alias.json()["emails"][0]["category"], "phishing")

        stats = self.client.get("/api/stats", params={"mailbox_id": mailbox.id})
        self.assertEqual(stats.status_code, 200, stats.text)
        self.assertEqual(stats.json()["categories"]["warn"], 1)
        self.assertEqual(stats.json()["categories"]["spam"], 1)
        self.assertEqual(stats.json()["categories"]["phishing"], 1)
        self.assertEqual(stats.json()["categories"]["malware"], 0)

    def test_fp_and_fn_both_create_truthful_pending_training_samples(self):
        false_positive_email = QuarantineEmail(
            email_id="training-false-positive",
            subject="Actually safe",
            label="WARN",
            category="warn",
            status="warn_delivered",
            fused_score=0.68,
            ml_probability=0.76,
            raw_content="From: sender@example.test\r\n\r\nSafe content",
            sender="sender@example.test",
            recipient_list="review@example.test",
        )
        false_negative_email = QuarantineEmail(
            email_id="training-false-negative",
            subject="Actually phishing",
            label="CLEAN",
            category="clean",
            status="released",
            fused_score=0.2,
            ml_probability=0.3,
            raw_content="From: attacker@example.test\r\n\r\nDangerous content",
            sender="attacker@example.test",
            recipient_list="review@example.test",
        )
        self.db.add_all([false_positive_email, false_negative_email])
        self.db.commit()

        fp_response = self.client.post(
            f"/api/emails/{false_positive_email.email_id}/report-false-positive",
            json={"notes": "Verified business message"},
        )
        self.assertEqual(fp_response.status_code, 200, fp_response.text)
        self.assertEqual(fp_response.json()["training_status"], "pending_review")

        fn_response = self.client.post(
            f"/api/emails/{false_negative_email.email_id}/report-false-negative",
            json={"corrected_label": "phishing", "notes": "Credential theft link"},
        )
        self.assertEqual(fn_response.status_code, 200, fn_response.text)
        self.assertEqual(fn_response.json()["status"], "pending_review")

        samples = self.db.query(TrainingSample).order_by(TrainingSample.email_id).all()
        self.assertEqual(len(samples), 2)
        by_id = {sample.email_id: sample for sample in samples}
        self.assertEqual(by_id[false_positive_email.email_id].feedback_type, "false_positive")
        self.assertEqual(by_id[false_positive_email.email_id].corrected_label, "clean")
        self.assertEqual(by_id[false_positive_email.email_id].status, "pending")
        self.assertEqual(by_id[false_negative_email.email_id].feedback_type, "false_negative")
        self.assertEqual(by_id[false_negative_email.email_id].corrected_label, "phishing")

        self.db.refresh(false_positive_email)
        self.db.refresh(false_negative_email)
        self.assertEqual((false_positive_email.label, false_positive_email.status), ("CLEAN", "released"))
        self.assertEqual(
            (false_negative_email.label, false_negative_email.category, false_negative_email.status),
            ("QUARANTINE", "phishing", "confirmed_threat"),
        )

        stats_response = self.client.get("/api/admin/training/stats")
        self.assertEqual(stats_response.status_code, 200, stats_response.text)
        training_stats = stats_response.json()
        self.assertEqual(training_stats["by_status"]["pending"], 2)
        self.assertEqual(training_stats["by_feedback_type"]["false_positive"], 1)
        self.assertEqual(training_stats["by_feedback_type"]["false_negative"], 1)
        self.assertFalse(training_stats["retraining"]["configured"])

    def test_trash_is_visible_only_to_assigned_admin_or_superadmin(self):
        admin = User(
            username="trash-test-admin",
            hashed_password=hash_password("Admin-password-123"),
            role="admin",
            is_active=True,
        )
        mailbox = AdminMailbox(
            email="trash-owner@example.test",
            domain="example.test",
            password_hash=hash_password("Mailbox-password-123"),
            sender_name="Trash Owner",
            assigned_to=admin.username,
            created_by=self.superadmin.username,
            is_active=True,
        )
        trashed = QuarantineEmail(
            email_id="admin-only-trash-message",
            subject="Administrative trash item",
            label="CLEAN",
            category="clean",
            status="trash",
            fused_score=0.0,
            sender="sender@example.test",
            recipient_list=mailbox.email,
        )
        self.db.add_all([admin, mailbox, trashed])
        self.db.commit()

        mailbox_client = TestClient(app, base_url="http://localhost")
        admin_client = TestClient(app, base_url="http://localhost")
        try:
            login = mailbox_client.post(
                "/api/auth/login",
                data={"username": mailbox.email, "password": "Mailbox-password-123"},
            )
            self.assertEqual(login.status_code, 200, login.text)
            self.assertEqual(
                mailbox_client.get("/api/emails", params={"folder": "trash"}).status_code,
                403,
            )
            self.assertEqual(
                mailbox_client.get(f"/api/emails/{trashed.email_id}").status_code,
                404,
            )
            self.assertEqual(
                mailbox_client.post(f"/api/emails/{trashed.email_id}/restore").status_code,
                403,
            )

            admin_token = create_access_token({"sub": admin.username, "role": admin.role})
            admin_client.cookies.set("access_token", admin_token)
            admin_trash = admin_client.get(
                "/api/emails",
                params={"mailbox_id": mailbox.id, "folder": "trash"},
            )
            self.assertEqual(admin_trash.status_code, 200, admin_trash.text)
            self.assertEqual(
                [row["email_id"] for row in admin_trash.json()["emails"]],
                [trashed.email_id],
            )
            restored = admin_client.post(f"/api/emails/{trashed.email_id}/restore")
            self.assertEqual(restored.status_code, 200, restored.text)
            self.assertEqual(restored.json()["status"], "released")

            moved_again = admin_client.delete(f"/api/emails/{trashed.email_id}")
            self.assertEqual(moved_again.status_code, 200, moved_again.text)
            self.assertEqual(moved_again.json()["status"], "trash")
            permanent = admin_client.delete(f"/api/emails/{trashed.email_id}")
            self.assertEqual(permanent.status_code, 200, permanent.text)
            self.assertEqual(permanent.json()["status"], "deleted")
        finally:
            mailbox_client.close()
            admin_client.close()

    def test_user_account_can_login_with_email_case_insensitively(self):
        user = User(
            username="email-login-user",
            email="person@example.test",
            hashed_password=hash_password("User-password-123"),
            role="user",
            is_active=True,
        )
        self.db.add(user)
        self.db.commit()

        client = TestClient(app, base_url="http://localhost")
        try:
            login = client.post(
                "/api/auth/login",
                data={"username": "PERSON@EXAMPLE.TEST", "password": "User-password-123"},
            )
            self.assertEqual(login.status_code, 200, login.text)
            self.assertEqual(login.json()["role"], "user")
        finally:
            client.close()

    def test_mailbox_reports_are_scoped_to_assigned_admin(self):
        admin_one = User(
            username="report-admin-one",
            email="admin-one@example.test",
            hashed_password=hash_password("Admin-one-password-123"),
            role="admin",
            is_active=True,
        )
        admin_two = User(
            username="report-admin-two",
            email="admin-two@example.test",
            hashed_password=hash_password("Admin-two-password-123"),
            role="admin",
            is_active=True,
        )
        mailbox = AdminMailbox(
            email="reports@example.test",
            domain="example.test",
            password_hash=hash_password("Reports-password-123"),
            sender_name="Reports",
            assigned_to=admin_one.username,
            created_by=admin_one.username,
            is_active=True,
        )
        self.db.add_all([admin_one, admin_two, mailbox])
        self.db.commit()

        mailbox_client = TestClient(app, base_url="http://localhost")
        admin_one_client = TestClient(app, base_url="http://localhost")
        admin_two_client = TestClient(app, base_url="http://localhost")
        try:
            mailbox_client.cookies.set("mailbox_token", create_access_token({
                "sub": f"mailbox:{mailbox.id}",
                "role": "mailbox",
                "mailbox_id": str(mailbox.id),
                "mailbox_email": mailbox.email,
            }))
            submitted = mailbox_client.post("/api/reports", json={
                "mailbox_email": mailbox.email,
                "subject": "Masalah inbox",
                "message": "Pesan tidak tampil",
                "category": "bug",
            })
            self.assertEqual(submitted.status_code, 200, submitted.text)
            own_history = mailbox_client.get("/api/reports")
            self.assertEqual(own_history.status_code, 200, own_history.text)
            self.assertEqual(len(own_history.json()), 1)
            self.assertEqual(own_history.json()[0]["status"], "open")

            for client, admin in (
                (admin_one_client, admin_one),
                (admin_two_client, admin_two),
            ):
                client.cookies.set("access_token", create_access_token({
                    "sub": admin.username,
                    "role": admin.role,
                }))

            visible = admin_one_client.get("/api/admin/reports")
            hidden = admin_two_client.get("/api/admin/reports")
            self.assertEqual(visible.status_code, 200, visible.text)
            self.assertEqual(hidden.status_code, 200, hidden.text)
            self.assertEqual(len(visible.json()), 1)
            self.assertEqual(visible.json()[0]["mailbox_email"], mailbox.email)
            self.assertEqual(hidden.json(), [])

            progress = admin_one_client.put(
                f"/api/admin/reports/{submitted.json()['id']}",
                json={"admin_reply": "Laporan sedang diperiksa"},
            )
            self.assertEqual(progress.status_code, 200, progress.text)
            updated_history = mailbox_client.get("/api/reports")
            self.assertEqual(updated_history.status_code, 200, updated_history.text)
            self.assertEqual(updated_history.json()[0]["status"], "in_progress")
            self.assertEqual(updated_history.json()[0]["admin_reply"], "Laporan sedang diperiksa")
        finally:
            mailbox_client.close()
            admin_one_client.close()
            admin_two_client.close()

    def test_snoozed_mailbox_page_does_not_crash(self):
        response = self.client.get("/api/emails", params={"folder": "snoozed"})
        self.assertEqual(response.status_code, 200, response.text)
        self.assertIn("emails", response.json())

    def test_autosave_updates_one_stable_draft_with_latest_content(self):
        draft_id = "draft_clientautosave123456"
        first = self.client.post("/api/emails/draft", json={
            "draft_id": draft_id,
            "from_email": "bantuan@example.test",
            "to": "recipient@example.test",
            "subject": "Versi awal",
            "body": "Isi awal",
        })
        self.assertEqual(first.status_code, 200, first.text)
        self.assertEqual(first.json()["email_id"], draft_id)

        latest = self.client.post("/api/emails/draft", json={
            "draft_id": draft_id,
            "from_email": "bantuan@example.test",
            "to": "recipient@example.test",
            "subject": "Versi terbaru",
            "body": "Isi terbaru yang harus dipulihkan",
        })
        self.assertEqual(latest.status_code, 200, latest.text)
        self.assertEqual(
            self.db.query(QuarantineEmail).filter(QuarantineEmail.email_id == draft_id).count(),
            1,
        )
        self.db.expire_all()
        saved = self.db.query(QuarantineEmail).filter(QuarantineEmail.email_id == draft_id).one()
        self.assertEqual(saved.subject, "Versi terbaru")
        self.assertEqual(saved.raw_content, "Isi terbaru yang harus dipulihkan")

    def test_read_status_updates_the_entire_reply_thread(self):
        messages = [
            QuarantineEmail(
                email_id="thread-incoming-1",
                subject="Status thread",
                label="CLEAN",
                status="released",
                fused_score=0.0,
                sender="Sender <sender@example.test>",
                recipient_list="mailbox@example.test",
                is_read=True,
                message_id_header="<thread-incoming-1@example.test>",
            ),
            QuarantineEmail(
                email_id="thread-sent-1",
                subject="Re: Status thread",
                label="SENT",
                status="sent",
                fused_score=0.0,
                sender="mailbox@example.test",
                recipient_list="Sender <sender@example.test>",
                is_read=False,
                message_id_header="<thread-sent-1@example.test>",
                references_header="<thread-incoming-1@example.test>",
            ),
            QuarantineEmail(
                email_id="thread-incoming-2",
                subject="Re: Status thread",
                label="CLEAN",
                status="released",
                fused_score=0.0,
                sender="Sender <sender@example.test>",
                recipient_list="mailbox@example.test",
                is_read=False,
                message_id_header="<thread-incoming-2@example.test>",
                references_header="<thread-incoming-1@example.test> <thread-sent-1@example.test>",
            ),
            QuarantineEmail(
                email_id="thread-incoming-3",
                subject="Re: Status thread",
                label="CLEAN",
                status="released",
                fused_score=0.0,
                sender="Sender <sender@example.test>",
                recipient_list="mailbox@example.test, observer@example.test",
                is_read=False,
                message_id_header="<thread-incoming-3@example.test>",
                references_header="<thread-incoming-2@example.test>",
            ),
        ]
        self.db.add_all(messages)
        self.db.commit()

        response = self.client.put("/api/emails/thread-incoming-1/read", json={"is_read": True})
        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(response.json()["updated_count"], 3)
        self.db.expire_all()
        incoming = self.db.query(QuarantineEmail).filter(
            QuarantineEmail.email_id.in_(["thread-incoming-1", "thread-incoming-2", "thread-incoming-3"])
        ).all()
        self.assertTrue(all(message.is_read for message in incoming))

        detail = self.client.get("/api/emails/thread-incoming-1")
        self.assertEqual(detail.status_code, 200, detail.text)
        self.assertTrue(detail.json()["thread_is_read"])
        self.assertFalse(detail.json()["thread_has_unread"])

        response = self.client.put("/api/emails/thread-incoming-2/read", json={"is_read": False})
        self.assertEqual(response.status_code, 200, response.text)
        self.db.expire_all()
        incoming = self.db.query(QuarantineEmail).filter(
            QuarantineEmail.email_id.in_(["thread-incoming-1", "thread-incoming-2", "thread-incoming-3"])
        ).all()
        self.assertTrue(all(not message.is_read for message in incoming))

    def test_superadmin_can_create_and_hard_delete_admin_by_username(self):
        with patch.dict(os.environ, {"VITE_MAIL_DOMAIN": "example.test"}):
            create_response = self.client.post(
                "/api/admin/users",
                json={
                    "username": "delete-flow-admin",
                    "email": "delete-flow-admin@example.test",
                    "password": "test-password-123",
                    "role": "admin",
                },
            )
        self.assertEqual(create_response.status_code, 201, create_response.text)
        created = self.db.query(User).filter(User.username == "delete-flow-admin").one()
        self.assertIsNone(created.email)
        self.assertNotIn("email", create_response.json())

        delete_response = self.client.delete("/api/admin/users/delete-flow-admin/hard")
        self.assertEqual(delete_response.status_code, 200, delete_response.text)
        self.assertIsNone(
            self.db.query(User).filter(User.username == "delete-flow-admin").first()
        )

    def test_profile_avatar_upload_and_profile_response(self):
        mailbox = AdminMailbox(
            email="avatar-mailbox@example.test",
            domain="example.test",
            password_hash=hash_password("Avatar-password-123"),
            sender_name="Avatar Mailbox",
            created_by=self.superadmin.username,
            is_active=True,
        )
        self.db.add(mailbox)
        self.db.commit()

        image_buffer = io.BytesIO()
        Image.new("RGB", (32, 32), color=(30, 100, 220)).save(image_buffer, format="PNG")
        image_buffer.seek(0)

        with tempfile.TemporaryDirectory() as temp_dir:
            with patch("dashboard.app.static_dir", Path(temp_dir)):
                upload_response = self.client.post(
                    "/api/auth/profile/avatar",
                    files={"avatar": ("avatar.png", image_buffer.getvalue(), "image/png")},
                )
                self.assertEqual(upload_response.status_code, 200, upload_response.text)
                avatar_url = upload_response.json()["avatar_url"]
                self.assertTrue(avatar_url.startswith("/static/avatars/user_"))
                self.assertTrue((Path(temp_dir) / "avatars" / Path(avatar_url).name).is_file())

                profile_response = self.client.get("/api/auth/profile")
                self.assertEqual(profile_response.status_code, 200, profile_response.text)
                self.assertEqual(profile_response.json()["avatar_url"], avatar_url)
                self.assertNotIn("email", profile_response.json())
                me_response = self.client.get("/api/auth/me")
                self.assertEqual(me_response.status_code, 200, me_response.text)
                self.assertEqual(me_response.json()["user"]["avatar_url"], avatar_url)

                mailbox_upload = self.client.post(
                    "/api/auth/profile/avatar",
                    params={"mailbox_id": mailbox.id},
                    files={"avatar": ("mailbox-avatar.png", image_buffer.getvalue(), "image/png")},
                )
                self.assertEqual(mailbox_upload.status_code, 200, mailbox_upload.text)
                mailbox_avatar_url = mailbox_upload.json()["avatar_url"]
                self.assertTrue(mailbox_avatar_url.startswith("/static/avatars/mailbox_"))
                self.assertTrue((Path(temp_dir) / "avatars" / Path(mailbox_avatar_url).name).is_file())

                mailbox_profile = self.client.get(
                    "/api/auth/profile",
                    params={"mailbox_id": mailbox.id},
                )
                self.assertEqual(mailbox_profile.status_code, 200, mailbox_profile.text)
                self.assertEqual(mailbox_profile.json()["avatar_url"], mailbox_avatar_url)

                mailbox_list = self.client.get("/api/admin/mailboxes")
                self.assertEqual(mailbox_list.status_code, 200, mailbox_list.text)
                listed_mailbox = next(
                    row for row in mailbox_list.json() if row["id"] == mailbox.id
                )
                self.assertEqual(listed_mailbox["avatar_url"], mailbox_avatar_url)

    def test_admin_profile_update_keeps_mailbox_ownership_and_refreshes_login(self):
        admin = User(
            username="profile-admin",
            email="profile-admin@example.test",
            hashed_password=hash_password("AdminPass123"),
            role="admin",
            is_active=True,
        )
        self.db.add(admin)
        self.db.flush()
        mailbox = AdminMailbox(
            email="managed@example.test",
            domain="example.test",
            assigned_to=admin.username,
            created_by=admin.username,
            is_active=True,
        )
        self.db.add(mailbox)
        self.db.flush()
        self.db.add(AdminMailboxAccess(mailbox_id=mailbox.id, username=admin.username))
        self.db.commit()

        admin_client = TestClient(app, base_url="http://localhost")
        try:
            admin_client.cookies.set(
                "access_token",
                create_access_token({"sub": admin.username, "role": admin.role}),
            )
            response = admin_client.put(
                "/api/auth/profile",
                json={
                    "username": "profile-admin-renamed",
                    "current_password": "AdminPass123",
                    "new_password": "NewAdminPass456",
                },
            )
            self.assertEqual(response.status_code, 200, response.text)

            self.db.expire_all()
            renamed = self.db.query(User).filter(User.username == "profile-admin-renamed").first()
            self.assertIsNotNone(renamed)
            mailbox = self.db.query(AdminMailbox).filter(AdminMailbox.email == "managed@example.test").one()
            self.assertEqual(mailbox.assigned_to, "profile-admin-renamed")
            self.assertEqual(mailbox.created_by, "profile-admin-renamed")
            access = self.db.query(AdminMailboxAccess).filter(AdminMailboxAccess.mailbox_id == mailbox.id).one()
            self.assertEqual(access.username, "profile-admin-renamed")

            me_response = admin_client.get("/api/auth/me")
            self.assertEqual(me_response.status_code, 200, me_response.text)
            self.assertEqual(me_response.json()["user"]["username"], "profile-admin-renamed")

            old_login = admin_client.post(
                "/api/auth/login",
                data={"username": "profile-admin-renamed", "password": "AdminPass123"},
            )
            self.assertEqual(old_login.status_code, 401, old_login.text)
            new_login = admin_client.post(
                "/api/auth/login",
                data={"username": "profile-admin-renamed", "password": "NewAdminPass456"},
            )
            self.assertEqual(new_login.status_code, 200, new_login.text)
        finally:
            admin_client.close()

    def test_detection_and_quarantine_are_scoped_to_real_assigned_mailboxes(self):
        password = "Review-password-123"
        admin_one = User(
            username="review-admin-one", email="review-one@example.test",
            hashed_password=hash_password(password), role="admin",
            organization_id=self.organization.id, is_active=True,
        )
        admin_two = User(
            username="review-admin-two", email="review-two@example.test",
            hashed_password=hash_password(password), role="admin",
            organization_id=self.organization.id, is_active=True,
        )
        mailbox_one = AdminMailbox(
            email="review-a@example.test", domain="example.test", sender_name="Review A",
            assigned_to=admin_one.username, created_by=self.superadmin.username, is_active=True,
        )
        mailbox_two = AdminMailbox(
            email="review-b@example.test", domain="example.test", sender_name="Review B",
            assigned_to=admin_two.username, created_by=self.superadmin.username, is_active=True,
        )
        own_clean = QuarantineEmail(
            email_id="review-own-clean", label="CLEAN", category="clean", status="released",
            subject="Real clean inbound", sender="safe@example.test",
            recipient_list="Review A <review-a@example.test>", fused_score=0.08,
        )
        own_threat = QuarantineEmail(
            email_id="review-own-threat", label="QUARANTINE", category="phishing", status="quarantined",
            subject="Real own threat", sender="attacker@example.test",
            recipient_list="Review A <review-a@example.test>", fused_score=0.94,
        )
        other_threat = QuarantineEmail(
            email_id="review-other-threat", label="QUARANTINE", category="spam", status="quarantined",
            subject="Other admin threat", sender="spam@example.test",
            recipient_list="review-b@example.test", fused_score=0.88,
        )
        draft = QuarantineEmail(
            email_id="review-draft", label="DRAFT", category="draft", status="draft",
            subject="Draft is not a detection", sender=mailbox_one.email,
            recipient_list=mailbox_one.email, fused_score=0,
        )
        self.db.add_all([
            admin_one, admin_two, mailbox_one, mailbox_two,
            own_clean, own_threat, other_threat, draft,
        ])
        self.db.commit()

        admin_client = TestClient(app, base_url="http://localhost")
        try:
            admin_client.cookies.set("access_token", create_access_token({
                "sub": admin_one.username,
                "role": admin_one.role,
            }))
            detection = admin_client.get("/api/admin/detection-logs")
            self.assertEqual(detection.status_code, 200, detection.text)
            detection_data = detection.json()
            self.assertEqual(
                {row["email_id"] for row in detection_data["logs"]},
                {own_clean.email_id, own_threat.email_id},
            )
            self.assertTrue(all(row["mailbox"] == mailbox_one.email for row in detection_data["logs"]))
            self.assertEqual([row["email"] for row in detection_data["mailboxes"]], [mailbox_one.email])

            quarantine = admin_client.get("/api/admin/quarantine")
            self.assertEqual(quarantine.status_code, 200, quarantine.text)
            self.assertEqual([row["email_id"] for row in quarantine.json()["emails"]], [own_threat.email_id])
            self.assertEqual(quarantine.json()["emails"][0]["mailbox"], mailbox_one.email)

            forbidden_filter = admin_client.get(
                "/api/admin/quarantine",
                params={"mailbox": mailbox_two.email},
            )
            self.assertEqual(forbidden_filter.status_code, 403, forbidden_filter.text)
            audit = admin_client.get("/api/audit-log")
            self.assertEqual(audit.status_code, 403, audit.text)
        finally:
            admin_client.close()

    def test_profile_update_rejects_invalid_username_and_weak_password(self):
        invalid_username = self.client.put(
            "/api/auth/profile",
            json={"username": "bad name", "current_password": "test-password-123"},
        )
        self.assertEqual(invalid_username.status_code, 400, invalid_username.text)

        weak_password = self.client.put(
            "/api/auth/profile",
            json={
                "username": self.superadmin.username,
                "current_password": "test-password-123",
                "new_password": "onlylowercase",
            },
        )
        self.assertEqual(weak_password.status_code, 400, weak_password.text)

    def test_profile_username_can_change_without_changing_password(self):
        original_password = "test-password-123"
        response = self.client.put(
            "/api/auth/profile",
            json={"username": "super-renamed-only"},
        )
        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(response.json()["username"], "super-renamed-only")

        me_response = self.client.get("/api/auth/me")
        self.assertEqual(me_response.status_code, 200, me_response.text)
        self.assertEqual(me_response.json()["user"]["username"], "super-renamed-only")

        login_response = self.client.post(
            "/api/auth/login",
            data={"username": "super-renamed-only", "password": original_password},
        )
        self.assertEqual(login_response.status_code, 200, login_response.text)

    def test_profile_password_change_still_requires_current_password(self):
        response = self.client.put(
            "/api/auth/profile",
            json={
                "username": self.superadmin.username,
                "new_password": "StrongPassword456",
            },
        )
        self.assertEqual(response.status_code, 400, response.text)
        self.assertIn("Current password is required", response.json()["detail"])

    def test_global_admin_cannot_manage_users_but_can_manage_domain_mailboxes(self):
        admin = User(
            username="domain-admin",
            email="admin@managed.test",
            hashed_password=hash_password("test-password-123"),
            role="admin",
            organization_id=None,
            is_active=True,
        )
        self.db.add(admin)
        self.db.commit()

        admin_client = TestClient(app, base_url="http://localhost")
        try:
            login_response = admin_client.post(
                "/api/auth/login",
                data={"username": admin.username, "password": "test-password-123"},
            )
            self.assertEqual(login_response.status_code, 200, login_response.text)

            with patch.dict(os.environ, {"VITE_MAIL_DOMAIN": "managed.test"}):
                config_response = admin_client.get("/api/admin/config")
                self.assertEqual(config_response.status_code, 200, config_response.text)
                self.assertEqual(config_response.json()["mail_domain"], "managed.test")

                create_user_response = admin_client.post(
                    "/api/admin/users",
                    json={
                        "username": "managed-user",
                        "email": "managed-user@managed.test",
                        "password": "test-password-123",
                        "role": "user",
                    },
                )
                self.assertEqual(create_user_response.status_code, 403, create_user_response.text)

                list_response = admin_client.get("/api/admin/users")
                self.assertEqual(list_response.status_code, 403, list_response.text)

                mailbox_response = admin_client.post(
                    "/api/admin/mailboxes",
                    json={
                        "email": "support@managed.test",
                        "domain": "managed.test",
                        "password": "Strong-Test-123!",
                        "sender_name": "Support",
                    },
                )
                self.assertEqual(mailbox_response.status_code, 200, mailbox_response.text)
                mailbox_id = self.db.query(AdminMailbox).filter(
                    AdminMailbox.email == "support@managed.test"
                ).one().id

                forward_response = admin_client.put(
                    f"/api/admin/mailboxes/{mailbox_id}/forwarder",
                    json={
                        "target": "archive@example.test",
                        "enabled": True,
                        "keep_copy": True,
                    },
                )
                self.assertEqual(forward_response.status_code, 200, forward_response.text)
                self.assertTrue(forward_response.json()["forward_enabled"])

                token_response = admin_client.post(
                    f"/api/admin/mailboxes/{mailbox_id}/autologin-token"
                )
                self.assertEqual(token_response.status_code, 200, token_response.text)
                self.assertTrue(token_response.json().get("token"))
        finally:
            admin_client.close()

    def test_admin_overview_stats_are_scoped_to_assigned_mailboxes(self):
        admin_one = User(
            username="overview-admin-one",
            email="overview-admin-one@example.test",
            hashed_password=hash_password("Admin-password-123"),
            role="admin",
            organization_id=self.organization.id,
            is_active=True,
        )
        admin_two = User(
            username="overview-admin-two",
            email="overview-admin-two@example.test",
            hashed_password=hash_password("Admin-password-123"),
            role="admin",
            organization_id=self.organization.id,
            is_active=True,
        )
        mailbox_user_one = User(
            username="overview-user-one",
            email="overview-one@example.test",
            hashed_password=hash_password("User-password-123"),
            role="user",
            organization_id=self.organization.id,
            is_active=True,
        )
        mailbox_user_two = User(
            username="overview-user-two",
            email="overview-two@example.test",
            hashed_password=hash_password("User-password-123"),
            role="user",
            organization_id=self.organization.id,
            is_active=True,
        )
        mailbox_one = AdminMailbox(
            email=mailbox_user_one.email,
            domain="example.test",
            password_hash=hash_password("Mailbox-password-123"),
            sender_name="Overview One",
            assigned_to=admin_one.username,
            created_by=self.superadmin.username,
            is_active=True,
        )
        mailbox_two = AdminMailbox(
            email=mailbox_user_two.email,
            domain="example.test",
            password_hash=hash_password("Mailbox-password-123"),
            sender_name="Overview Two",
            assigned_to=admin_two.username,
            created_by=self.superadmin.username,
            is_active=True,
        )
        self.db.add_all([
            admin_one,
            admin_two,
            mailbox_user_one,
            mailbox_user_two,
            mailbox_one,
            mailbox_two,
            QuarantineEmail(
                email_id="overview-one-clean",
                label="CLEAN",
                category="clean",
                fused_score=0.02,
                status="released",
                sender="sender-one@example.test",
                recipient_list=mailbox_one.email,
            ),
            QuarantineEmail(
                email_id="overview-two-phishing",
                label="QUARANTINE",
                category="phishing",
                fused_score=0.95,
                status="quarantined",
                sender="sender-two@example.test",
                recipient_list=mailbox_two.email,
            ),
        ])
        self.db.commit()

        def scoped_client(user):
            client = TestClient(app, base_url="http://localhost")
            client.cookies.set("access_token", create_access_token({
                "sub": user.username,
                "role": user.role,
            }))
            return client

        client_one = scoped_client(admin_one)
        client_two = scoped_client(admin_two)
        try:
            stats_one = client_one.get("/api/admin/stats")
            stats_two = client_two.get("/api/admin/stats")
            self.assertEqual(stats_one.status_code, 200, stats_one.text)
            self.assertEqual(stats_two.status_code, 200, stats_two.text)
            self.assertEqual(stats_one.json()["total_mailboxes"], 1)
            self.assertEqual(stats_one.json()["active_mailboxes"], 1)
            self.assertEqual(stats_one.json()["total_regular_users"], 1)
            self.assertEqual(stats_one.json()["total_emails"], 1)
            self.assertEqual(stats_one.json()["clean"], 1)
            self.assertEqual(stats_one.json()["categories"]["phishing"], 0)
            self.assertEqual(stats_two.json()["total_emails"], 1)
            self.assertEqual(stats_two.json()["clean"], 0)
            self.assertEqual(stats_two.json()["categories"]["phishing"], 1)

            super_stats = self.client.get("/api/admin/stats")
            self.assertEqual(super_stats.status_code, 200, super_stats.text)
            self.assertEqual(super_stats.json()["total_mailboxes"], 2)
            self.assertEqual(super_stats.json()["active_mailboxes"], 2)
            self.assertEqual(super_stats.json()["total_regular_users"], 2)
            self.assertEqual(super_stats.json()["total_emails"], 2)
        finally:
            client_one.close()
            client_two.close()

    def test_mailbox_listing_login_and_empty_inbox(self):
        mailbox = AdminMailbox(
            email="route-inbox@example.test",
            domain="example.test",
            password_hash=hash_password("mailbox-password-123"),
            sender_name="Route Inbox",
            created_by=self.superadmin.username,
            is_active=True,
        )
        self.db.add(mailbox)
        self.db.flush()
        self.db.add(
            AdminMailboxAccess(
                mailbox_id=mailbox.id,
                username=self.superadmin.username,
            )
        )
        self.db.commit()

        list_response = self.client.get("/api/user/mailboxes")
        self.assertEqual(list_response.status_code, 200, list_response.text)
        self.assertIn(mailbox.id, [item["id"] for item in list_response.json()])

        token_response = self.client.post(
            f"/api/admin/mailboxes/{mailbox.id}/autologin-token"
        )
        self.assertEqual(token_response.status_code, 200, token_response.text)
        autologin_client = TestClient(app, base_url="http://localhost")
        try:
            redeem_response = autologin_client.post(
                "/api/mailboxes/autologin",
                json={"token": token_response.json()["token"]},
            )
            self.assertEqual(redeem_response.status_code, 200, redeem_response.text)
            me_response = autologin_client.get("/api/auth/me")
            self.assertEqual(me_response.status_code, 200, me_response.text)
            self.assertTrue(me_response.json()["authenticated"])
            self.assertEqual(me_response.json()["user"]["role"], "mailbox")
            autologin_inbox = autologin_client.get(
                "/api/emails",
                params={"mailbox_id": mailbox.id, "mailbox": mailbox.email},
            )
            self.assertEqual(autologin_inbox.status_code, 200, autologin_inbox.text)
        finally:
            autologin_client.close()

        mailbox_client = TestClient(app, base_url="http://localhost")
        try:
            login_response = mailbox_client.post(
                "/api/mailboxes/login",
                json={
                    "email": mailbox.email,
                    "password": "mailbox-password-123",
                },
            )
            self.assertEqual(login_response.status_code, 200, login_response.text)

            profile_response = mailbox_client.get("/api/auth/profile")
            self.assertEqual(profile_response.status_code, 200, profile_response.text)
            self.assertEqual(profile_response.json()["mailbox_email"], mailbox.email)

            inbox_response = mailbox_client.get("/api/emails")
            self.assertEqual(inbox_response.status_code, 200, inbox_response.text)
        finally:
            mailbox_client.close()

    def test_mailbox_status_and_permanent_delete_are_distinct_and_atomic(self):
        mailbox = AdminMailbox(
            email="delete-me@managed.test",
            domain="managed.test",
            password_hash=hash_password("Strong-Test-123!"),
            sender_name="Delete Me",
            assigned_to="mailbox-manager",
            created_by=self.superadmin.username,
            is_active=True,
        )
        manager = User(
            username="mailbox-manager",
            email="manager@managed.test",
            hashed_password=hash_password("Strong-Test-123!"),
            role="admin",
            is_active=True,
        )
        incoming = QuarantineEmail(
            email_id="delete-incoming",
            label="CLEAN",
            fused_score=0.0,
            status="released",
            sender="Sender <sender@example.test>",
            recipient_list="Delete Me <delete-me@managed.test>",
        )
        outgoing = QuarantineEmail(
            email_id="delete-outgoing",
            label="SENT",
            fused_score=0.0,
            status="sent",
            sender="Delete Me <delete-me@managed.test>",
            recipient_list="recipient@example.test",
        )
        unrelated = QuarantineEmail(
            email_id="keep-unrelated",
            label="CLEAN",
            fused_score=0.0,
            status="released",
            sender="sender@example.test",
            recipient_list="other@managed.test",
        )
        self.db.add_all([manager, mailbox, incoming, outgoing, unrelated])
        self.db.flush()
        self.db.add_all([
            AdminMailboxAccess(mailbox_id=mailbox.id, username=manager.username),
            Feedback(email_id=incoming.email_id, feedback_type="correct", notes="delete"),
            TrainingSample(
                email_id=outgoing.email_id,
                raw_email="raw",
                original_label="CLEAN",
                corrected_label="spam",
                feedback_type="relabel",
                reported_by=manager.username,
            ),
            AuditLog(user=manager.username, action="read", email_id=incoming.email_id),
            AuditTrail(
                actor=manager.username,
                action="inference",
                target_type="email",
                target_id=outgoing.email_id,
                status="SUCCESS",
            ),
        ])
        self.db.commit()
        mailbox_id = mailbox.id
        incoming_id = incoming.email_id
        outgoing_id = outgoing.email_id
        unrelated_id = unrelated.email_id

        access_before = self.client.get(f"/api/mailboxes/{mailbox_id}/access")
        self.assertEqual(access_before.status_code, 200, access_before.text)

        token_response = self.client.post(
            f"/api/admin/mailboxes/{mailbox_id}/autologin-token"
        )
        self.assertEqual(token_response.status_code, 200, token_response.text)
        mailbox_client = TestClient(app, base_url="http://localhost")
        redeem = mailbox_client.post(
            "/api/mailboxes/autologin", json={"token": token_response.json()["token"]}
        )
        self.assertEqual(redeem.status_code, 200, redeem.text)
        mailbox_me = mailbox_client.get("/api/auth/me")
        self.assertTrue(mailbox_me.json()["authenticated"])

        disable = self.client.put(
            f"/api/admin/mailboxes/{mailbox_id}", json={"is_active": False}
        )
        self.assertEqual(disable.status_code, 200, disable.text)
        self.assertFalse(disable.json()["is_active"])
        self.assertEqual(
            self.db.query(AdminMailboxAccess).filter_by(mailbox_id=mailbox_id).count(), 1
        )
        self.assertEqual(self.db.query(QuarantineEmail).count(), 3)
        disabled_access = self.client.get(f"/api/mailboxes/{mailbox_id}/access")
        self.assertEqual(disabled_access.status_code, 404, disabled_access.text)

        activate = self.client.put(
            f"/api/admin/mailboxes/{mailbox_id}", json={"is_active": True}
        )
        self.assertEqual(activate.status_code, 200, activate.text)
        self.assertTrue(activate.json()["is_active"])
        self.assertEqual(
            self.db.query(AdminMailboxAccess).filter_by(mailbox_id=mailbox_id).count(), 1
        )
        active_access = self.client.get(f"/api/mailboxes/{mailbox_id}/access")
        self.assertEqual(active_access.status_code, 200, active_access.text)

        permanent_delete = self.client.delete(f"/api/admin/mailboxes/{mailbox_id}")
        self.assertEqual(permanent_delete.status_code, 200, permanent_delete.text)
        self.assertEqual(permanent_delete.json()["deleted"]["emails"], 2)
        self.db.expire_all()
        self.assertIsNone(self.db.query(AdminMailbox).filter_by(id=mailbox_id).first())
        self.assertEqual(
            [row.email_id for row in self.db.query(QuarantineEmail).all()],
            [unrelated_id],
        )
        self.assertEqual(self.db.query(Feedback).filter_by(email_id=incoming_id).count(), 0)
        self.assertEqual(self.db.query(TrainingSample).filter_by(email_id=outgoing_id).count(), 0)
        self.assertEqual(self.db.query(AuditLog).filter_by(email_id=incoming_id).count(), 0)
        self.assertEqual(self.db.query(AuditTrail).filter_by(target_id=outgoing_id).count(), 0)
        self.assertEqual(self.db.query(AdminMailboxAccess).filter_by(mailbox_id=mailbox_id).count(), 0)
        deleted_access = self.client.get(f"/api/mailboxes/{mailbox_id}/access")
        self.assertEqual(deleted_access.status_code, 404, deleted_access.text)
        deleted_inbox = self.client.get(
            "/api/emails", params={"mailbox_id": mailbox_id, "folder": "inbox"}
        )
        self.assertEqual(deleted_inbox.status_code, 404, deleted_inbox.text)
        deleted_mailbox_me = mailbox_client.get("/api/auth/me")
        self.assertFalse(deleted_mailbox_me.json()["authenticated"])
        mailbox_client.close()

    def test_admin_mailbox_ownership_is_strict_and_reassignment_is_atomic(self):
        password = "Strong-Test-123!"
        admin_one = User(
            username="mail-admin-one",
            email="admin-one@managed.test",
            hashed_password=hash_password(password),
            role="admin",
            is_active=True,
        )
        admin_two = User(
            username="mail-admin-two",
            email="admin-two@managed.test",
            hashed_password=hash_password(password),
            role="admin",
            is_active=True,
        )
        mailbox_a = AdminMailbox(
            email="a@managed.test",
            domain="managed.test",
            password_hash=hash_password(password),
            sender_name="Mailbox A",
            assigned_to=admin_one.username,
            created_by=self.superadmin.username,
            is_active=True,
        )
        mailbox_b = AdminMailbox(
            email="b@managed.test",
            domain="managed.test",
            password_hash=hash_password(password),
            sender_name="Mailbox B",
            assigned_to=admin_two.username,
            created_by=self.superadmin.username,
            is_active=True,
        )
        self.db.add_all([admin_one, admin_two, mailbox_a, mailbox_b])
        self.db.flush()
        self.db.add_all([
            AdminMailboxAccess(mailbox_id=mailbox_a.id, username=admin_one.username),
            AdminMailboxAccess(mailbox_id=mailbox_b.id, username=admin_two.username),
            QuarantineEmail(
                email_id="mailbox-a-draft-1",
                label="DRAFT",
                fused_score=0.0,
                status="draft",
                sender=mailbox_a.email,
                recipient_list="recipient@example.test",
            ),
            QuarantineEmail(
                email_id="mailbox-a-draft-2",
                label="DRAFT",
                fused_score=0.0,
                status="draft",
                sender=mailbox_a.email,
                recipient_list="recipient@example.test",
            ),
            QuarantineEmail(
                email_id="mailbox-b-clean-1",
                label="CLEAN",
                fused_score=0.0,
                status="released",
                sender="sender@example.test",
                recipient_list=mailbox_b.email,
            ),
        ])
        self.db.commit()

        def login_admin(username):
            client = TestClient(app, base_url="http://localhost")
            response = client.post(
                "/api/auth/login",
                data={"username": username, "password": password},
            )
            self.assertEqual(response.status_code, 200, response.text)
            return client

        client_one = login_admin(admin_one.username)
        client_two = login_admin(admin_two.username)
        try:
            list_one = client_one.get("/api/admin/mailboxes")
            list_two = client_two.get("/api/admin/mailboxes")
            self.assertEqual([row["email"] for row in list_one.json()], [mailbox_a.email])
            self.assertEqual([row["email"] for row in list_two.json()], [mailbox_b.email])
            self.assertEqual(list_one.json()[0]["assigned_to"], admin_one.username)

            stats_a = client_one.get("/api/stats", params={"mailbox_id": mailbox_a.id})
            stats_b = client_two.get("/api/stats", params={"mailbox_id": mailbox_b.id})
            self.assertEqual(stats_a.status_code, 200, stats_a.text)
            self.assertEqual(stats_b.status_code, 200, stats_b.text)
            self.assertEqual(stats_a.json()["draft"], 2)
            self.assertEqual(stats_a.json()["total"], 0)
            self.assertEqual(stats_b.json()["draft"], 0)
            self.assertEqual(stats_b.json()["total"], 1)

            denied_token = client_one.post(
                f"/api/admin/mailboxes/{mailbox_b.id}/autologin-token"
            )
            self.assertEqual(denied_token.status_code, 403, denied_token.text)
            denied_inbox = client_one.get(
                "/api/emails",
                params={"mailbox_id": mailbox_b.id, "mailbox": mailbox_b.email},
            )
            self.assertEqual(denied_inbox.status_code, 403, denied_inbox.text)
            allowed_token = client_two.post(
                f"/api/admin/mailboxes/{mailbox_b.id}/autologin-token"
            )
            self.assertEqual(allowed_token.status_code, 200, allowed_token.text)

            reassign = self.client.put(
                f"/api/admin/mailboxes/{mailbox_b.id}",
                json={"assigned_to": admin_one.username},
            )
            self.assertEqual(reassign.status_code, 200, reassign.text)
            self.assertEqual(reassign.json()["assigned_to"], admin_one.username)

            list_one_after = client_one.get("/api/admin/mailboxes")
            list_two_after = client_two.get("/api/admin/mailboxes")
            self.assertEqual(
                {row["email"] for row in list_one_after.json()},
                {mailbox_a.email, mailbox_b.email},
            )
            self.assertEqual(list_two_after.json(), [])

            manager_access = self.db.query(AdminMailboxAccess).filter(
                AdminMailboxAccess.mailbox_id == mailbox_b.id,
                AdminMailboxAccess.username.in_([admin_one.username, admin_two.username]),
            ).all()
            self.assertEqual([row.username for row in manager_access], [admin_one.username])

            delete_owner = self.client.delete(f"/api/admin/users/{admin_one.username}/hard")
            self.assertEqual(delete_owner.status_code, 409, delete_owner.text)
        finally:
            client_one.close()
            client_two.close()


    def test_comprehensive_reports_use_live_scoped_email_rows(self):
        manager = User(
            username="report-manager",
            email="report-manager@example.test",
            hashed_password=hash_password("Report-password-123"),
            role="admin",
            is_active=True,
        )
        mailbox = AdminMailbox(
            email="security-report@example.test",
            domain="example.test",
            password_hash=hash_password("Mailbox-password-123"),
            sender_name="Security Report",
            assigned_to=manager.username,
            created_by=self.superadmin.username,
            is_active=True,
        )
        other_mailbox = AdminMailbox(
            email="other-security-report@example.test",
            domain="example.test",
            password_hash=hash_password("Mailbox-password-123"),
            sender_name="Other Security Report",
            assigned_to=manager.username,
            created_by=self.superadmin.username,
            is_active=True,
        )
        inbound = QuarantineEmail(
            email_id="real-report-email-001",
            label="QUARANTINE",
            category="phishing",
            fused_score=0.9234,
            ml_probability=0.8877,
            sa_score=7.5,
            anomaly_score=0.2,
            status="quarantined",
            subject="URGENT real database subject with enough text to verify PDF wrapping",
            sender="Attacker <attacker@example.test>",
            recipient_list="Security Team <security-report@example.test>",
            spf_result="FAIL",
            dkim_result="FAIL",
            dmarc_result="FAIL",
            routing_reason="DMARC failed and fused score exceeded the quarantine threshold",
            model_version="report-test-v1",
        )
        sent = QuarantineEmail(
            email_id="sent-must-not-be-reported",
            label="SENT",
            fused_score=0.0,
            status="sent",
            subject="Not an inbound detection",
            sender=mailbox.email,
            recipient_list="external@example.test",
        )
        unrelated = QuarantineEmail(
            email_id="other-mailbox-email",
            label="CLEAN",
            fused_score=0.1,
            status="released",
            subject="Outside selected mailbox",
            sender="sender@example.test",
            recipient_list="unmanaged@example.test",
        )
        other_inbound = QuarantineEmail(
            email_id="other-selected-mailbox-email",
            label="CLEAN",
            category="clean",
            fused_score=0.05,
            status="released",
            subject="Must be excluded by mailbox selection",
            sender="sender@example.test",
            recipient_list=other_mailbox.email,
        )
        self.db.add_all([manager, mailbox, other_mailbox, inbound, sent, unrelated, other_inbound])
        self.db.commit()

        payload = {
            "format": "excel",
            "admin_ids": [manager.id],
            "mailbox_ids": [mailbox.id],
            "include_users": True,
            "include_emails": True,
        }
        excel_response = self.client.post("/api/admin/export/generate", json=payload)
        self.assertEqual(excel_response.status_code, 200, excel_response.text)
        self.assertEqual(
            excel_response.headers["content-type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        with zipfile.ZipFile(io.BytesIO(excel_response.content)) as workbook:
            workbook_text = "\n".join(
                workbook.read(name).decode("utf-8", errors="ignore")
                for name in workbook.namelist()
                if name.endswith(".xml")
            )
        self.assertIn(inbound.email_id, workbook_text)
        self.assertIn(mailbox.email, workbook_text)
        self.assertIn(manager.username, workbook_text)
        self.assertIn("report-test-v1", workbook_text)
        self.assertNotIn(sent.email_id, workbook_text)
        self.assertNotIn(unrelated.email_id, workbook_text)
        self.assertNotIn(other_inbound.email_id, workbook_text)
        self.assertNotIn(other_mailbox.email, workbook_text)
        selected_workbook = load_workbook(io.BytesIO(excel_response.content), read_only=True)
        mailbox_sheets = [name for name in selected_workbook.sheetnames if name.startswith("Mail ")]
        self.assertEqual(len(mailbox_sheets), 1)
        selected_sheet = selected_workbook[mailbox_sheets[0]]
        self.assertEqual(selected_sheet["B1"].value, mailbox.email)
        self.assertEqual(selected_sheet["B2"].value, manager.username)

        payload["format"] = "csv"
        csv_response = self.client.post("/api/admin/export/generate", json=payload)
        self.assertEqual(csv_response.status_code, 400, csv_response.text)
        self.assertIn("pdf or excel", csv_response.json()["detail"])

        payload["format"] = "pdf"
        pdf_response = self.client.post("/api/admin/export/generate", json=payload)
        self.assertEqual(pdf_response.status_code, 200, pdf_response.text)
        self.assertEqual(pdf_response.headers["content-type"], "application/pdf")
        self.assertTrue(pdf_response.content.startswith(b"%PDF"))
        self.assertGreater(len(pdf_response.content), 2000)
        pdf_text = "\n".join(
            page.extract_text() or ""
            for page in PdfReader(io.BytesIO(pdf_response.content)).pages
        )
        self.assertIn(f"Mailbox Report: {mailbox.email}", pdf_text)
        self.assertIn(inbound.subject, re.sub(r"\s+", " ", pdf_text))
        self.assertNotIn(other_mailbox.email, pdf_text)

    def test_email_analytics_uses_real_mailboxes_and_enforces_admin_scope(self):
        password = "Analytics-password-123"
        admin_one = User(
            username="analytics-admin-one", email="analytics-one@example.test",
            hashed_password=hash_password(password), role="admin", is_active=True,
        )
        admin_two = User(
            username="analytics-admin-two", email="analytics-two@example.test",
            hashed_password=hash_password(password), role="admin", is_active=True,
        )
        mailbox_one = AdminMailbox(
            email="analytics-a@example.test", domain="example.test",
            password_hash=hash_password(password), sender_name="Analytics A",
            assigned_to=admin_one.username, created_by=self.superadmin.username, is_active=True,
        )
        mailbox_two = AdminMailbox(
            email="analytics-b@example.test", domain="example.test",
            password_hash=hash_password(password), sender_name="Analytics B",
            assigned_to=admin_two.username, created_by=admin_two.username, is_active=True,
        )
        real_threat = QuarantineEmail(
            email_id="analytics-real-threat", label="QUARANTINE", category="phishing",
            fused_score=0.91, ml_probability=0.88, sa_score=6.2, status="quarantined",
            subject="Real analytics threat", sender="sender@example.test",
            recipient_list="Analytics A <analytics-a@example.test>",
            spf_result="FAIL", dkim_result="PASS", dmarc_result="FAIL",
        )
        sent_row = QuarantineEmail(
            email_id="analytics-sent-row", label="SENT", fused_score=0,
            status="sent", sender=mailbox_one.email, recipient_list="outside@example.test",
        )
        other_mailbox = QuarantineEmail(
            email_id="analytics-other-mailbox", label="CLEAN", fused_score=0.05,
            status="released", sender="safe@example.test", recipient_list=mailbox_two.email,
        )
        self.db.add_all([admin_one, admin_two, mailbox_one, mailbox_two, real_threat, sent_row, other_mailbox])
        self.db.commit()

        super_response = self.client.get("/api/admin/email-analytics")
        self.assertEqual(super_response.status_code, 200, super_response.text)
        super_rows = {row["email"]: row for row in super_response.json()}
        self.assertEqual(set(super_rows), {mailbox_one.email, mailbox_two.email})
        self.assertEqual(super_rows[mailbox_one.email]["admin"], admin_one.username)
        self.assertEqual(super_rows[mailbox_one.email]["total_emails"], 1)
        self.assertEqual(super_rows[mailbox_one.email]["phishing"], 1)
        self.assertEqual(super_rows[mailbox_two.email]["total_emails"], 1)

        detail = self.client.get(f"/api/admin/email-analytics/{mailbox_one.id}")
        self.assertEqual(detail.status_code, 200, detail.text)
        self.assertEqual([row["email_id"] for row in detail.json()["recent_threats"]], [real_threat.email_id])

        admin_client = TestClient(app, base_url="http://localhost")
        try:
            admin_client.cookies.set("access_token", create_access_token({
                "sub": admin_one.username,
                "role": admin_one.role,
            }))
            own_rows = admin_client.get("/api/admin/email-analytics")
            self.assertEqual(own_rows.status_code, 200, own_rows.text)
            self.assertEqual([row["email"] for row in own_rows.json()], [mailbox_one.email])
            forbidden = admin_client.get(f"/api/admin/email-analytics/{mailbox_two.id}")
            self.assertEqual(forbidden.status_code, 403, forbidden.text)
        finally:
            admin_client.close()

    def test_regular_user_cannot_reference_or_report_another_mailbox_email(self):
        user = User(
            username="limited-user",
            email="limited@example.test",
            hashed_password=hash_password("Limited-password-123"),
            role="user",
            is_active=True,
        )
        victim_email = QuarantineEmail(
            email_id="private-victim-email",
            label="CLEAN",
            category="clean",
            status="released",
            fused_score=0.05,
            subject="Private victim content",
            sender="sender@example.test",
            recipient_list="victim@example.test",
            raw_content="private body that must never be forwarded",
        )
        self.db.add_all([user, victim_email])
        self.db.commit()

        user_client = TestClient(app, base_url="http://localhost")
        try:
            user_client.cookies.set("access_token", create_access_token({
                "sub": user.username,
                "role": user.role,
            }))
            false_negative = user_client.post(
                f"/api/emails/{victim_email.email_id}/report-false-negative",
                json={"corrected_label": "phishing", "notes": "not mine"},
            )
            self.assertEqual(false_negative.status_code, 403, false_negative.text)

            for action, recipient_field in (
                ("reply", {"to": "external@example.test"}),
                ("forward", {"to": "external@example.test"}),
                ("share", {"share_with": "external@example.test"}),
            ):
                response = user_client.post(
                    "/api/emails/send",
                    json={
                        **recipient_field,
                        "from_email": user.email,
                        "subject": "Access check",
                        "body": "Must be rejected before SMTP",
                        "reply_to_id": victim_email.email_id,
                        "action": action,
                    },
                )
                self.assertEqual(response.status_code, 403, response.text)
        finally:
            user_client.close()

    def test_admin_threat_breakdown_is_limited_to_assigned_mailboxes(self):
        admin = User(
            username="scoped-breakdown-admin",
            email="scoped-admin@example.test",
            hashed_password=hash_password("Scoped-password-123"),
            role="admin",
            is_active=True,
        )
        own_mailbox = AdminMailbox(
            email="scoped-breakdown@example.test",
            domain="example.test",
            assigned_to=admin.username,
            created_by=self.superadmin.username,
            is_active=True,
        )
        own_threat = QuarantineEmail(
            email_id="scoped-breakdown-own",
            label="QUARANTINE",
            category="phishing",
            status="quarantined",
            fused_score=0.9,
            subject="Own threat",
            sender="attacker@example.test",
            recipient_list=own_mailbox.email,
        )
        foreign_threat = QuarantineEmail(
            email_id="scoped-breakdown-foreign",
            label="QUARANTINE",
            category="spam",
            status="quarantined",
            fused_score=0.9,
            subject="Foreign threat",
            sender="spam@example.test",
            recipient_list="foreign@example.test",
        )
        self.db.add_all([admin, own_mailbox, own_threat, foreign_threat])
        self.db.commit()

        admin_client = TestClient(app, base_url="http://localhost")
        try:
            admin_client.cookies.set("access_token", create_access_token({
                "sub": admin.username,
                "role": admin.role,
            }))
            response = admin_client.get("/api/admin/threat-breakdown", params={"days": 7})
            self.assertEqual(response.status_code, 200, response.text)
            data = response.json()
            self.assertEqual(data["category_counts"]["phishing"], 1)
            self.assertEqual(data["category_counts"]["spam"], 0)
            self.assertEqual(
                [row["recipient"] for row in data["top_recipients"]],
                [own_mailbox.email],
            )
        finally:
            admin_client.close()


if __name__ == "__main__":
    unittest.main()
